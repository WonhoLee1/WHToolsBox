import mujoco
import mujoco.viewer
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import os
import json
import time
import shutil
from datetime import datetime
from run_discrete_builder import create_model, get_default_config
import pickle
from dataclasses import dataclass, field
from typing import Any, Dict, List
import math
import io

@dataclass
class DropSimResult:
    """시뮬레이션 전체 결과 데이터를 담는 통합 클래스 (최적화 모델 매칭용)"""
    config: Dict[str, Any]
    metrics: Dict[str, Any]
    max_g_force: float
    time_history: List[float]
    z_hist: List[float]
    root_acc_history: List[Any]
    corner_acc_hist: List[Any]
    
    pos_hist: List[Any] = field(default_factory=list)
    vel_hist: List[Any] = field(default_factory=list)
    acc_hist: List[Any] = field(default_factory=list)
    
    cog_pos_hist: List[Any] = field(default_factory=list)
    cog_vel_hist: List[Any] = field(default_factory=list)
    cog_acc_hist: List[Any] = field(default_factory=list)
    
    corner_pos_hist: List[Any] = field(default_factory=list)
    corner_vel_hist: List[Any] = field(default_factory=list)
    
    ground_impact_hist: List[float] = field(default_factory=list)
    air_drag_hist: List[float] = field(default_factory=list)
    air_viscous_hist: List[float] = field(default_factory=list)
    air_squeeze_hist: List[float] = field(default_factory=list)
    
    def save(self, filepath: str):
        with open(filepath, "wb") as f:
            pickle.dump(self, f)
            
    @classmethod
    def load(cls, filepath: str):
        with open(filepath, "rb") as f:
            return pickle.load(f)

# =====================================================================
# [NEW] 전역 로깅 설정 (모든 터미널 출력을 리포트 파일에 저장)
# =====================================================================
_report_file_path = None

def log_and_print(msg):
    """
    터미널에 출력함과 동시에, 현재 진행 중인 시뮬레이션의 
    summary_report.txt 파일이 설정되어 있다면 해당 파일에도 기록합니다.
    """
    content = str(msg)
    print(content) # 실제 터미널 출력
    
    global _report_file_path
    if _report_file_path:
        try:
            # 'a' (append) 모드로 열어 파일 끝에 추가 기록
            with open(_report_file_path, "a", encoding="utf-8") as rf:
                rf.write(content + "\n")
        except Exception as e:
            # 로깅 자체의 에러는 터미널에 최소한으로 출력
            print(f"!!! [Logging Failed] {e}")

# =====================================================================
# [NEW] 리포팅 및 설정을 위한 유틸리티 함수
# =====================================================================
def format_config_report(config, timestamp):
    """설정(Config) 데이터를 보기 좋게 문자열로 정리하여 반환합니다."""
    lines = []
    lines.append("\n" + "="*90)
    lines.append(f"  [MuJoCo Drop Simulation - Configuration Summary]  -  {timestamp}")
    lines.append("="*90)
    
    # 주요 카테고리별 정렬 정의
    categories = {
        "1. 낙하 조건 (Drop Setup)":       ["drop_mode", "drop_height", "sim_duration"],
        "2. 구조물 및 조립 (Assembly)":    ["include_paperbox", "include_cushion", "box_use_weld", "cush_use_weld", "oc_use_weld", "chas_use_weld"],
        "3. 형상 파라미터 (Geometry)":    ["box_w", "box_h", "box_d", "box_thick", "box_div", "cush_div", "assy_div", "oc_div", "chassis_div"],
        "4. 질량/재질 설정 (Material)":   ["mass_paper", "mass_cushion", "mass_oc", "mass_occ", "mass_chassis", "ground_friction"],
        "5. 솔버/환경 설정 (Solver)":     ["sim_integrator", "sim_timestep", "sim_iterations", "sim_impratio", "sim_nthread", "ground_solref", "tape_solref"],
        "6. 공기 역학 (Aerodynamics)":    ["enable_air_drag", "enable_air_squeeze", "air_density", "air_coef_squeeze", "air_cd_drag"]
    }
    
    logged_keys = set()
    for cat, keys in categories.items():
        lines.append(f"\n[{cat}]")
        for k in keys:
            if k in config:
                val = config[k]
                lines.append(f"  - {k:<25}: {val}")
                logged_keys.add(k)
    
    # 추가 질량(Aux Masses) 별도 처리
    if "chassis_aux_masses" in config and config["chassis_aux_masses"]:
        lines.append("\n[7. 추가 질량 (Chassis Aux Masses)]")
        for i, aux in enumerate(config["chassis_aux_masses"]):
            name = aux.get('name', f'Aux_{i}')
            lines.append(f"  * {name:<23}: pos={aux.get('pos')}, mass={aux.get('mass')}kg, size={aux.get('size')}")
        logged_keys.add("chassis_aux_masses")

    # 상기 카테고리에 누락된 기타 설정들
    other_keys = [k for k in config.keys() if k not in logged_keys and not k.startswith("mat_")]
    if other_keys:
        lines.append("\n[8. 기타 세부 설정 (Others)]")
        for k in sorted(other_keys):
            lines.append(f"  - {k:<25}: {config[k]}")
        
    lines.append("\n" + "="*90 + "\n")
    return "\n".join(lines)

def print_inertia_report(config, title="[Assembly Inertia Report]", logger=print):
    """지정된 설정에 따른 모델의 질량, CoG, MoI 정보를 측정하고 리포트 형식으로 출력합니다."""
    # 실제 XML 파일 생성을 피하기 위해 임시 파일명을 사용하지만, 
    # create_model 내부에서 이미 로깅 로직이 있으므로 이를 활용합니다.
    temp_xml = "temp_inertia_check.xml"
    _, total_mass, cog, moi, details = create_model(temp_xml, config=config, logger=logger)
    return total_mass, cog, moi, details

def save_config_as_py(config, filepath, timestamp):
    """전체 설정을 차후 재현 가능한 Python 딕셔너리 형태의 파일로 저장합니다."""
    import pprint
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f'# MuJoCo Drop Simulation Reproduction Config\n')
        f.write(f'# Generated for Run: rds-{timestamp}\n')
        f.write(f'# Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
        formatted_cfg = pprint.pformat(config, indent=4, sort_dicts=False)
        f.write(f'config = {formatted_cfg}\n')

plt.rcParams.update({'font.size': 8})



def ask_use_viewer():
    """사용자에게 GUI 뷰어(Passive Viewer) 사용 여부를 묻습니다."""
    try:
        prompt = "\n>> MuJoCo GUI 뷰어를 실행하시겠습니까? (Y/n, 기본값=Y): "
        ans = input(prompt).strip().lower()
        
        # [LOGGING] 사용자 응답을 로그에 기록
        log_and_print(f"{prompt.strip()} -> {ans if ans else 'y (default)'}")
        
        if ans == 'n':
            return False
        return True
    except EOFError:
        log_and_print(">> IO Error or Empty input. Defaulting to True.")
        return True


def calc_mujoco_stiffness(solref_str, mass=1.0):
    """
    MuJoCo solref (timeconst, dampratio)를 기반으로 유효 강성 K와 감쇠 C를 계산합니다.
    MuJoCo 공식: 
      K = 1 / (timeconst^2 * dampratio^2)
      C = 2 / timeconst
    물질의 질량(mass)을 곱하여 최종 차원(N/m, Ns/m)을 반환합니다.
    """
    try:
        parts = [float(x) for x in str(solref_str).split()]
        if len(parts) < 2:
            return 0.0, 0.0
        tc, dr = parts[0], parts[1]
        
        # tc 또는 dr 이 0이면 수치적 발산 방지
        if tc <= 0 or dr <= 0:
            return 0.0, 0.0
            
        k_unit = 1.0 / (tc**2 * dr**2)
        c_unit = 2.0 / tc
        
        return k_unit * mass, c_unit * mass
    except Exception:
        return 0.0, 0.0



def apply_aerodynamics(model, data, config):
    """
    공기 저항(Drag, Viscous) 및 지면 근접 시 압축 공기 효과(Squeeze Film)를 계산하여 적용합니다.
    [개선] 특정 한 면만 계산하는 방식에서 모든 하향 면(Downward Faces)을 체크하는 방식으로 변경하여 
    Corner/Edge 낙하 시에도 공기 쿠션 효과가 누락되지 않도록 합니다.
    """
    rho      = config.get('air_density', 1.225)
    mu       = config.get('air_viscosity', 1.81e-5)
    Cd_blunt = config.get('air_cd_drag', 1.05)
    Cd_visc  = config.get('air_cd_viscous', 0.0)
    Csq      = config.get('air_coef_squeeze', 0.0)
    h_sq_max = config.get('air_squeeze_hmax', 0.20)
    h_sq_min = config.get('air_squeeze_hmin', 0.001)
    
    # 적용 대상 바디 (최상위 박스)
    body_name = "BPackagingBox"
    body_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, body_name)
    if body_id == -1:
        return 0.0, 0.0, 0.0

    # 외력 초기화 (MuJoCo는 제어 콜백에서 직접 xfrc_applied를 설정하면 해당 스텝에 반영됨)
    data.xfrc_applied[body_id, :] = 0.0

    # CoM 상태 정보 추출
    pos      = data.xpos[body_id]
    mat      = data.xmat[body_id].reshape(3, 3)
    ang_vel  = data.cvel[body_id, 0:3] 
    lin_vel  = data.cvel[body_id, 3:6] 

    f_sq_total = 0.0
    total_torque = np.zeros(3)

    # [1] Squeeze Film 효과 계산 (모든 면에 대해)
    if config.get('enable_air_squeeze', True) and Csq > 0:
        dims = [config.get('box_w', 2.0), config.get('box_h', 1.4), config.get('box_d', 0.25)]
        # 각 로컬 축(X, Y, Z)에 대하여
        for axis_idx in range(3):
            axis_vec = mat[:, axis_idx]  # 로컬 축의 월드 벡터
            dot_z = axis_vec[2]          # 월드 Z축 투영 성분
            
            if abs(dot_z) < 1e-3: continue # 지면과 수직인 면은 스퀴즈 없음
            
            # 하향 방향의 법선 설정
            local_norm = np.zeros(3)
            local_norm[axis_idx] = -np.sign(dot_z) * (dims[axis_idx] / 2.0)
            
            # 해당 면의 가로/세로 길이 결정
            other_indices = [i for i in range(3) if i != axis_idx]
            u_len = dims[other_indices[0]]
            v_len = dims[other_indices[1]]
            u_vec = mat[:, other_indices[0]]
            v_vec = mat[:, other_indices[1]]
            
            # 형상 계수 계산
            geo_factor = ((u_len * v_len) / (2 * (u_len + v_len))) ** 2
            P_COEF = 0.5 * rho * geo_factor * Csq
            
            face_center_rel = mat @ local_norm
            
            # 격자 적산 (N x N)
            N = 6 # 성능을 위해 약간 조정
            dA = (u_len * v_len) / (N * N)
            grid_steps = np.linspace(-0.5+0.5/N, 0.5-0.5/N, N)
            
            for u in grid_steps:
                for v in grid_steps:
                    rel_p = face_center_rel + (u * u_len) * u_vec + (v * v_len) * v_vec
                    h = pos[2] + rel_p[2]
                    
                    if h_sq_min < h < h_sq_max:
                        # 점 속도 계산
                        p_vel = lin_vel + np.cross(ang_vel, rel_p)
                        vz = p_vel[2]
                        if vz < 0:
                            dF = P_COEF * dA * (vz / h)**2
                            dF = min(dF, 500.0 / (N*N/64)) # 포인트당 캡
                            f_sq_total += dF
                            total_torque += np.cross(rel_p, np.array([0, 0, dF]))

        # 물리 엔진에 힘 적용
        data.xfrc_applied[body_id, 2] += f_sq_total
        data.xfrc_applied[body_id, 3:6] += total_torque

    # [2] 로깅용 에어로다이나믹 추정값 (Drag/Viscous)
    # MuJoCo가 계산하는 값을 근사하여 반환 (시각화용)
    total_area = 2*(dims[0]*dims[1] + dims[1]*dims[2] + dims[2]*dims[0]) if 'dims' in locals() else 5.0
    v_mag = np.linalg.norm(lin_vel)
    f_drag_est = 0.5 * rho * v_mag**2 * Cd_blunt * (total_area/6.0) if config.get('enable_air_drag', True) else 0.0
    f_visc_est = mu * v_mag * Cd_visc * total_area if config.get('enable_air_drag', True) else 0.0

    return f_drag_est, f_visc_est, f_sq_total



def get_body_kinematics(model, data, body_name):
    body_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, body_name)
    if body_id == -1:
        return None
    pos = data.xpos[body_id].copy()
    mat = data.xmat[body_id].reshape(3, 3).copy()
    vel = data.cvel[body_id].copy() # [w_x, w_y, w_z, v_x, v_y, v_z]
    acc = data.cacc[body_id].copy() # [alpha_x, alpha_y, alpha_z, a_x, a_y, a_z]
    return pos, mat, vel, acc

def compute_corners(center_pos, center_mat, box_w, box_h, box_d):
    """지정된 중심점과 회전 행렬을 기반으로 8개 모서리 좌표 역계산"""
    corners_local = []
    for x in [-box_w/2, box_w/2]:
        for y in [-box_h/2, box_h/2]:
            for z in [-box_d/2, box_d/2]:
                corners_local.append(np.array([x, y, z]))
    
    corners_global = []
    for loc in corners_local:
        glob = center_pos + center_mat @ loc
        corners_global.append(glob)
    return np.array(corners_global)

def compute_corner_kinematics(center_pos, center_mat, center_vel, center_acc, box_w, box_h, box_d):
    # center_vel: [wx, wy, wz, vx, vy, vz]
    w = center_vel[0:3]
    v = center_vel[3:6]
    
    # center_acc: [alphax, alphay, alphaz, ax, ay, az]
    alpha = center_acc[0:3]
    a = center_acc[3:6]
    
    corners_local = []
    for x in [-box_w/2, box_w/2]:
        for y in [-box_h/2, box_h/2]:
            for z in [-box_d/2, box_d/2]:
                corners_local.append(np.array([x, y, z]))
                
    results = []
    for loc in corners_local:
        # global offset vector from center
        r = center_mat @ loc
        
        # velocity = v + w x r
        v_corner = v + np.cross(w, r)
        
        # acceleration = a + alpha x r + w x (w x r)
        a_corner = a + np.cross(alpha, r) + np.cross(w, np.cross(w, r))
        
        results.append({
            'pos': center_pos + r,
            'vel': v_corner,
            'acc': a_corner
        })
    return results

def run_simulation(config_or_path, sim_duration=0.5):
    # [NEW] 결과 저장을 위한 타임스탬프 및 출력 경로 초기화 (가장 먼저 수행)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"rds-{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    
    # 전역 로깅 경로 설정
    global _report_file_path
    _report_file_path = os.path.abspath(os.path.join(output_dir, "summary_report.txt"))

    # 시작 시점에 빈 파일을 생성(Overwrite)하여 기록 준비
    with open(_report_file_path, "w", encoding="utf-8") as f:
        pass 

    if isinstance(config_or_path, str):
        # XML 경로가 직접 전달된 경우
        xml_path = config_or_path
        log_and_print(f"\nLoading MuJoCo model from XML: {xml_path}")
        model = mujoco.MjModel.from_xml_path(xml_path)
        config = get_default_config() # fallback
    else:
        # Config 딕셔너리가 전달된 경우 (모델 생성 필요)
        config = get_default_config(config_or_path)
        
        # [NEW] 로깅 디렉토리 생성 직후, 보정 전(Baseline) 관성 상태를 측정하여 요약에 포함할 준비
        baseline_stats = None
        if not config.get("only_generate_xml", False):
            # 보정 전 상태를 확인하기 위해 aux masses가 없는 임시 설정을 만듭니다.
            base_cfg = config.copy()
            if "chassis_aux_masses" in base_cfg:
                base_cfg["chassis_aux_masses"] = []
            
            # 로그 파일이 생성되었으므로, 여기에 Baseline 기록 (터미널 출력은 생략하거나 별도 처리)
            _, b_m, b_c, b_i, _ = create_model("temp_baseline.xml", config=base_cfg, logger=lambda x: None)
            baseline_stats = (b_m, b_c, b_i)

        # Detailed summary of current configuration
        summary_text = format_config_report(config, timestamp)
        
        # [NEW] Baseline 정보를 텍스트 하단에 병합
        if baseline_stats:
            b_m, b_c, b_i = baseline_stats
            extra = []
            extra.append("\n[Baseline Inertia (Pre-Correction)]")
            extra.append(f"  - Baseline Mass            : {b_m:12.4f} kg")
            extra.append(f"  - Baseline CoG             : ({b_c[0]:.4f}, {b_c[1]:.4f}, {b_c[2]:.4f})")
            extra.append(f"  - Baseline MoI (x,y,z)     : ({b_i[0]:.6f}, {b_i[1]:.6f}, {b_i[2]:.6f})")
            summary_text = summary_text.replace("\n" + "="*90 + "\n", "\n".join(extra) + "\n\n" + "="*90 + "\n")

        log_and_print(summary_text)

        config_py_path = os.path.join(output_dir, f"rds-{timestamp}_config.py")
        save_config_as_py(config, config_py_path, timestamp)
        log_and_print(f"  >> Reproducible config script saved to: {os.path.basename(config_py_path)}")

        log_and_print("\nGenerating discrete box model from config...")
        xml_file = "temp_drop_sim.xml"
        xml_str, *_ = create_model(xml_file, config=config, logger=log_and_print)
        xml_abs_path = os.path.abspath(xml_file)
        log_and_print(f"  >> Generated XML: {xml_abs_path}")
        
        model = mujoco.MjModel.from_xml_string(xml_str)
        sim_duration = config.get('sim_duration', sim_duration)
        xml_path = xml_abs_path

    # [NEW] 모델 생성만 수행하고 시뮬레이션은 건너뛰는 옵션 처리
    if config.get("only_generate_xml", False):
        if xml_path and os.path.exists(xml_path):
            shutil.copy(xml_path, os.path.join(output_dir, os.path.basename(xml_path)))
            log_and_print(f"  >> [XML Mode] XML saved to {output_dir}. Skipping simulation.")
        return None
    
    data = mujoco.MjData(model)
    
    # [INFO] MuJoCo 3.0+ 멀티코어 연산은 XML의 <option npoolthread="N"> 설정을 통해 활성화됩니다.
    nthread = config.get("sim_nthread", 4)
    if nthread > 1:
        log_and_print(f"  >> Multicore requested: {nthread} threads (via npoolthread).")
    
    # Identify bodies for structural metrics
    # Group by component name -> dict of (i,j,k) -> body_id
    components = {}
    nominal_local_pos = {}
    
    for i in range(model.nbody):
        name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_BODY, i)
        if name and name.startswith("b_"):
            parts = name.split("_")
            # 최소 'b_{comp}_{i}_{j}_{k}' 형식을 갖추기 위해 5개 이상의 토큰이 필요함
            if len(parts) >= 5:
                # 인덱스 (i, j, k)는 마지막 3개 세그먼트에서 추출
                try:
                    idx_i = int(parts[-3])
                    idx_j = int(parts[-2])
                    idx_k = int(parts[-1])
                    
                    # 컴포넌트 이름은 첫 'b_'와 인덱스 사이의 모든 세그먼트를 결합
                    comp_name = "_".join(parts[1:-3])
                    
                    if comp_name not in components:
                        components[comp_name] = {}
                    components[comp_name][(idx_i, idx_j, idx_k)] = i
                    nominal_local_pos[i] = model.body_pos[i].copy()
                except ValueError:
                    # 인덱스 변환 실패 시 (이산 블록 형식이 아님) 건너뜀
                    continue

    # Time setup
    dt = model.opt.timestep
    duration = config.get("sim_duration", 1.0)
    steps = int(duration / dt)
    
    time_history = []
    z_hist = []
    pos_hist = []
    vel_hist = []
    acc_hist = []
    cog_pos_hist = []
    cog_vel_hist = []
    cog_acc_hist = []
    corner_pos_hist = []
    corner_vel_hist = []
    corner_acc_hist = []
    ground_impact_hist = []
    air_drag_hist = []
    air_viscous_hist = []
    air_squeeze_hist = []
    
    # Metric histories per component by row (j) and individual blocks
    # [NEW] MuJoCo 물리 파라미터 (Stiffness K, Damping C) 사전 계산 및 출력
    cush_tc = config.get("cush_solref_stiff", 0.02)
    cush_dr = config.get("cush_solref_damp", 1.0)
    cush_solref = f"{cush_tc} {cush_dr}"
    
    # 쿠션 블록 한 개의 질량 (에너지 계산용)
    cush_div = config.get("cush_div", [5, 4, 3])
    num_cush_blocks = np.prod(cush_div)
    m_cush_block = config.get("mass_cushion", 1.0) / num_cush_blocks
    
    k_cush, c_cush = calc_mujoco_stiffness(cush_solref, m_cush_block)
    
    # [NEW] 탄성 계수 (Young's Modulus, E) 근사 계산
    # E = (K * L) / A  (K: 강성, L: 블록 두께, A: 단면적)
    cush_w = config.get("box_w", 2.0) - 2*config.get("box_thick", 0.01)
    cush_h = config.get("box_h", 1.4) - 2*config.get("box_thick", 0.01)
    cush_d = config.get("box_d", 0.25) - 2*config.get("box_thick", 0.01)

    avg_dx = cush_w / cush_div[0]
    avg_dy = cush_h / cush_div[1]
    avg_dz = cush_d / cush_div[2]
    
    # 단면적 (xy 평면 기준) 및 두께 (z 방향)
    area_avg = avg_dx * avg_dy
    E_estimated = (k_cush * avg_dz) / area_avg if area_avg > 0 else 0.0
    E_mpa = E_estimated / 1e6 # Pa -> MPa
    E_kpa = E_estimated / 1e3 # Pa -> kPa

    # 에너지 계산용 프록시 강성을 실제 물리값으로 업데이트
    k_spring_proxy = k_cush

    # 바닥 및 테이프 정보 (참조용 전역 질량 기준)
    total_mass = float(np.sum(model.body_mass))
    k_ground, c_ground = calc_mujoco_stiffness(config.get("ground_solref", "0.01 1.0"), total_mass)
    k_tape,   c_tape   = calc_mujoco_stiffness(config.get("tape_solref", "0.005 1.0"), config.get("mass_occ", 0.1))

    log_and_print("\n" + "="*80)
    log_and_print(f"[MuJoCo Solver Parameters (Calculated K & C)]")
    log_and_print(f" - Ground (Global) : K = {k_ground:10.1f} N/m, C = {c_ground:8.1f} Ns/m")
    log_and_print(f"   (Used: mass={total_mass:.3f}kg, solref={config.get('ground_solref', '0.01 1.0')})")
    
    log_and_print(f" - Tape (Global)   : K = {k_tape:10.1f} N/m, C = {c_tape:8.1f} Ns/m")
    log_and_print(f"   (Used: mass={config.get('mass_occ', 0.1):.3f}kg, solref={config.get('tape_solref', '0.005 1.0')})")
    
    log_and_print(f" - Cushion (Block) : K = {k_cush:10.1f} N/m, C = {c_cush:8.1f} Ns/m (per block)")
    log_and_print(f"   (Used: block_mass={m_cush_block:.6f}kg, solref={cush_solref})")
    log_and_print(f"   (Geom: Avg_Area={area_avg:.6f} m^2, Avg_Depth={avg_dz:.4f} m)")
    log_and_print(f"   >> Est. Young's Modulus (E): {E_mpa:.4e} MPa ({E_kpa:.2f} kPa)")
    log_and_print("="*80 + "\n")

    metrics = {}
    for comp in components:
        metrics[comp] = {}
        metrics[comp]['all_blocks_angle'] = {}
        metrics[comp]['block_nominals'] = {}
        for block_idx in components[comp]:
            metrics[comp]['all_blocks_angle'][block_idx] = []
            metrics[comp]['block_nominals'][block_idx] = nominal_local_pos[components[comp][block_idx]]
            
        # Find unique j indices
        j_idx = set([k[1] for k in components[comp].keys()])
        for j in j_idx:
            metrics[comp][j] = {
                'bending': [], 'twist': [], 'energy': [],
                'loc_b': [], 'loc_t': [], 'loc_e': []
            }
            
    log_and_print(f"Starting simulation for {duration} seconds with {steps} steps...")
    
    # k_spring_proxy is now calculated from MuJoCo parameters above
    
    prev_vel_z = 0.0
    
    mujoco.mj_forward(model, data)
    
    # [NEW] 영구 변형(Plastic Deformation) 로직 초기화
    geom_state_tracker = {}
    enable_plasticity = config.get("enable_plasticity", False)
    plasticity_ratio = config.get("plasticity_ratio", 0.3)
    
    def apply_plastic_deformation():
        if not enable_plasticity: return
        current_penetrations = {}
        
        # 1. 런타임에 바닥(ground)과 쿠션 상호작용 검사
        for i in range(data.ncon):
            con = data.contact[i]
            g1_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, con.geom1)
            g2_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, con.geom2)
            
            target_geom = -1
            if g1_name and "ground" in g1_name.lower(): target_geom = con.geom2
            elif g2_name and "ground" in g2_name.lower(): target_geom = con.geom1
            
            if target_geom != -1:
                t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, target_geom)
                if t_name and t_name.startswith("g_bcushion_"):
                    # g_bcushion_{i}_{j}_{k} 정규식 또는 스플릿 추출
                    parts = t_name.split('_')
                    if len(parts) >= 5:
                        try:
                            c_i = int(parts[2])
                            c_j = int(parts[3])
                            # 8곳의 모서리 및 두께 방향 능(Edge Columns)에 걸치는 패드만 허용
                            # i가 0 또는 끝이고, j가 0 또는 끝인 기둥들 (Z인 k는 무관)
                            if (c_i == 0 or c_i == cush_div[0] - 1) and (c_j == 0 or c_j == cush_div[1] - 1):
                                pen = -con.dist
                                if pen > 1e-4:
                                    if pen > current_penetrations.get(target_geom, 0.0):
                                        current_penetrations[target_geom] = pen
                        except ValueError:
                            pass

        # 2. 신규 접촉 등록
        for gid in current_penetrations:
            if gid not in geom_state_tracker:
                geom_state_tracker[gid] = {'max_p': 0.0, 'applied': False}
                
        # 3. 압축(Compression) 및 회복(Recovery) 상태 업데이트
        for gid, state in geom_state_tracker.items():
            curr_p = current_penetrations.get(gid, 0.0)
            
            # 압축 진행 중
            if curr_p >= state['max_p']:
                state['max_p'] = curr_p
                state['applied'] = False
            
            # 회복 중 변형 적용
            if state['max_p'] > 0.001 and not state['applied']:
                recovery = state['max_p'] - curr_p
                if recovery >= state['max_p'] * plasticity_ratio:
                    # 향후 탄성 계수 기반 공식을 위해 별도 분리 가능한 지점 
                    deformation = state['max_p'] * plasticity_ratio
                    
                    local_pos = model.geom_pos[gid]
                    inward_dir = -np.sign(local_pos[:3])
                    
                    # 비율: 80% 이동, 20% 축소 (옆 간섭 방지)
                    shrink = deformation * 0.2
                    shift = deformation * 0.8
                    
                    # 1D/2D 기준 크기 축소 
                    model.geom_size[gid][0] = max(0.001, model.geom_size[gid][0] - shrink)
                    model.geom_size[gid][1] = max(0.001, model.geom_size[gid][1] - shrink)
                    
                    # 위치 이동
                    model.geom_pos[gid] += inward_dir * shift
                    
                    # [NEW] 영구 변형량(shrink) 직관적 확인을 위한 색상 변경 (화이트 -> 블루)
                    # 수 mm 이하 변형 시에도 눈에 확 띄도록 2000.0을 곱해 색상을 바꿉니다.
                    color_drop = min(1.0, shrink * 2000.0)  
                    model.geom_rgba[gid][0] = max(0.0, 1.0 - color_drop)
                    model.geom_rgba[gid][1] = max(0.0, 1.0 - color_drop)
                    model.geom_rgba[gid][2] = 1.0  # Blue 고정
                    model.geom_rgba[gid][3] = 1.0  # Alpha 유지
                    
                    state['applied'] = True
                    # 로깅 추가 (작동 확인 가능하도록 주석 해제)
                    t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, gid)
                    print(f"  [Plasticity] {t_name} 영구 변형 발생! shrink={shrink*1000:.3f}mm, shift={shift*1000:.3f}mm")
            
            if curr_p == 0.0 and state['applied']:
                state['max_p'] = 0.0
                state['applied'] = False

    # [NEW] Control Callback 등록 (mjcb_control)
    def mjcb_aerodynamics(model, data):
        apply_aerodynamics(model, data, config)
    mujoco.set_mjcb_control(mjcb_aerodynamics)
    
    # [NEW] Keyboard Control State & Callback (Reference: boxmotionsim_v0_0_2)
    class ControlState:
        def __init__(self):
            self.paused = False
            self.step_request = False
            self.reset_request = False
            self.quit_request = False
    
    ctrl = ControlState()
    
    def key_callback(keycode):
        if keycode == 32: # Space
            ctrl.paused = not ctrl.paused
            log_and_print(f"   >> [VIEWER] {'PAUSED' if ctrl.paused else 'RUNNING'}")
        elif keycode == 262: # Right Arrow
            ctrl.step_request = True
        elif keycode == 256: # ESC
            ctrl.quit_request = True
            log_and_print("   >> [VIEWER] Quit requested.")
        elif keycode == 259 or keycode == 82: # Backspace (259) or R (82)
            ctrl.reset_request = True
    
    use_viewer = config.get("use_viewer", False)
    viewer = None
    if use_viewer:
        viewer = mujoco.viewer.launch_passive(model, data, key_callback=key_callback)
        log_and_print("\n" + "-"*40)
        log_and_print("🎮 [Passive Viewer Interactive Controls]")
        log_and_print(" - [Space]     : Pause / Resume")
        log_and_print(" - [Right]     : Single Step (when paused)")
        log_and_print(" - [Backspace] : Reset Simulation")
        log_and_print(" - [Esc]       : Stop & Show Results")
        log_and_print("-"*40 + "\n")

    # [NEW] 시뮬레이션 진행 상황 테이블 헤더 출력
    log_and_print("-" * 65)
    log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
    log_and_print("-" * 65)

    start_real_time = time.time()
    last_reported_interval = -1
    last_report_step = 0
    last_report_real_time = start_real_time
    report_count = 0
    
    step = 0
    while step < steps:
        if ctrl.quit_request:
            break
            
        if ctrl.reset_request:
            log_and_print("   >> [VIEWER] Resetting simulation...")
            mujoco.mj_resetData(model, data)
            # Re-initialize histories
            time_history.clear(); z_hist.clear(); pos_hist.clear(); vel_hist.clear(); acc_hist.clear()
            cog_pos_hist.clear(); cog_vel_hist.clear(); cog_acc_hist.clear()
            ground_impact_hist.clear(); air_drag_hist.clear(); air_viscous_hist.clear(); air_squeeze_hist.clear()
            corner_pos_hist.clear(); corner_vel_hist.clear(); corner_acc_hist.clear()
            for c in metrics:
                for k in metrics[c]:
                    if isinstance(metrics[c][k], dict):
                        for b in metrics[c][k]: 
                            if isinstance(metrics[c][k][b], list):
                                metrics[c][k][b].clear()
                    elif isinstance(metrics[c][k], list):
                        metrics[c][k].clear()
            step = 0
            ctrl.reset_request = False
            ctrl.paused = True
            
            # Reset table reporting
            last_reported_interval = -1
            start_real_time = time.time()
            last_report_step = 0
            last_report_real_time = start_real_time
            report_count = 0
            log_and_print("-" * 65)
            log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
            log_and_print("-" * 65)
            continue

        if not ctrl.paused or ctrl.step_request:
            try:
                mujoco.mj_step(model, data)
                apply_plastic_deformation()  # 매 스텝마다 변형 체킹
                ctrl.step_request = False
                step += 1
            except Exception as e:
                import sys
                sys.stderr.write(f"\n[Drop Sim Error] Simulation unstable: {e}\n")
                break
        else:
            # Paused state: Just sync viewer and sleep a bit to save CPU
            if viewer:
                viewer.sync()
            time.sleep(0.01)
            continue

        if viewer and viewer.is_running():
            viewer.sync()
            
        # [INFO] Progress Reporting (Table Row)
        sim_interval = int(data.time / 0.05)
        if sim_interval > last_reported_interval:
            # Interval FPS calculation
            current_real_time = time.time()
            real_elapsed = current_real_time - start_real_time
            interval_elapsed = current_real_time - last_report_real_time
            if interval_elapsed > 0:
                fps = (step - last_report_step) / interval_elapsed
            else:
                fps = 0.0
            
            log_and_print(f"| {step:8d} | {data.time:12.3f} | {real_elapsed:13.2f} | {fps:8.1f} |")
            
            last_reported_interval = sim_interval
            last_report_step = step
            last_report_real_time = current_real_time
            report_count += 1
            
            # [NEW] 30개 출력마다 열 제목 재출력
            if report_count > 0 and report_count % 30 == 0:
                log_and_print("-" * 65)
                log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
                log_and_print("-" * 65)
            
        time_history.append(data.time)
        
        # 1. Root Kinematics (BPackagingBox itself is just a massless container)
        pos, mat, vel, acc = get_body_kinematics(model, data, "BPackagingBox")
        z_hist.append(pos[2])
        pos_hist.append(pos)
        vel_hist.append(vel)
        acc_hist.append(acc)
        
        root_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, "BPackagingBox")
        if root_id != -1:
            cog_pos_hist.append(data.subtree_com[root_id].copy())
            cog_vel_hist.append(data.subtree_linvel[root_id].copy())
        else:
            cog_pos_hist.append(pos)
            cog_vel_hist.append(vel[3:6])
        cog_acc_hist.append(acc[3:6]) # Subtree acceleration in MuJoCo requires numerical differentiation, so proxy by root acceleration
        
        corners = compute_corner_kinematics(pos, mat, vel, acc, config["box_w"], config["box_h"], config["box_d"])
        corner_pos_hist.append([c['pos'] for c in corners])
        corner_vel_hist.append([c['vel'] for c in corners])
        corner_acc_hist.append([c['acc'] for c in corners])
        
        # Ground Impact (Sum of normal forces from worldbody contacts)
        ground_f = 0.0
        for i_con in range(data.ncon):
            contact = data.contact[i_con]
            body1 = model.geom_bodyid[contact.geom1]
            body2 = model.geom_bodyid[contact.geom2]
            if body1 == 0 or body2 == 0:
                forces = np.zeros(6, dtype=np.float64)
                mujoco.mj_contactForce(model, data, i_con, forces)
                ground_f += abs(forces[0])
        ground_impact_hist.append(ground_f)
        
        # Air Resistance Logging (Already applied via callback)
        f_drag, f_viscous, f_squeeze = apply_aerodynamics(model, data, config)
        
        air_drag_hist.append(f_drag)
        air_viscous_hist.append(f_viscous)
        air_squeeze_hist.append(f_squeeze)
        
        # G-Force is now calculated purely from root kinematic differences after simulation
            
        # 2. Component Structural Metrics
        mat_inv = mat.T
        for comp, blocks in components.items():
            # Gather current local positions
            current_local = {}
            for (curr_i, curr_j, curr_k), b_id in blocks.items():
                g_pos = data.xpos[b_id]
                g_mat = data.xmat[b_id].reshape(3, 3)
                rel_pos = g_pos - pos
                l_pos = mat_inv @ rel_pos
                current_local[(curr_i, curr_j, curr_k)] = l_pos
                
                # Calculate angular deformation (angle of rotation relative to root)
                rot_rel = mat_inv @ g_mat
                tr = np.clip(np.trace(rot_rel), -1.0, 3.0)
                theta = np.arccos((tr - 1.0) / 2.0)
                block_angle_deg = np.degrees(theta)
                metrics[comp]['all_blocks_angle'][(curr_i, curr_j, curr_k)].append(block_angle_deg)
                
            # Process by row (j)
            j_indices = sorted(list(set([k[1] for k in blocks.keys()])))
            j_mid = j_indices[len(j_indices)//2] if len(j_indices) > 0 else 0
            
            # Mid row slope for twist reference
            mid_row_x = []
            mid_row_z = []
            for (curr_i, curr_j, curr_k), l_pos in current_local.items():
                if curr_j == j_mid:
                    mid_row_x.append(l_pos[0])
                    mid_row_z.append(l_pos[2])
            mid_slope = 0.0
            if len(mid_row_x) > 1:
                mid_slope = np.polyfit(mid_row_x, mid_row_z, 1)[0]
                
            for j in j_indices:
                row_x = []
                row_z = []
                energies = {}
                dz_vals = {}
                
                for (curr_i, curr_j, curr_k), l_pos in current_local.items():
                    if curr_j == j:
                        b_id = blocks[(curr_i, curr_j, curr_k)]
                        nom_pos = nominal_local_pos[b_id]
                        delta = l_pos - nom_pos
                        
                        # Energy proxy: 0.5 * k * dx^2
                        energy = 0.5 * k_spring_proxy * np.sum(delta**2)
                        energies[(curr_i, curr_k)] = energy
                        
                        row_x.append(l_pos[0])
                        row_z.append(l_pos[2])
                        dz_vals[(curr_i, curr_k)] = delta[2]
                
                # Bending: Max Z deflection angle approximation
                bending_deg = 0.0
                loc_b = "N/A"
                if len(dz_vals) > 1:
                    max_idx = max(dz_vals, key=dz_vals.get)
                    min_idx = min(dz_vals, key=dz_vals.get)
                    dz_diff = dz_vals[max_idx] - dz_vals[min_idx]
                    dx_diff = max(row_x) - min(row_x) if max(row_x) != min(row_x) else 1e-6
                    bending_deg = np.degrees(np.arctan(abs(dz_diff / dx_diff)))
                    loc_b = f"{max_idx[0]}_{j}_{max_idx[1]}"
                    
                # Twisting: relative slope diff to mid row
                twist_deg = 0.0
                loc_t = f"*_{j}_*"
                if len(row_x) > 1:
                    slope = np.polyfit(row_x, row_z, 1)[0]
                    twist_deg = np.degrees(np.arctan(abs(slope - mid_slope)))
                    
                # Energy Peak
                peak_energy = 0.0
                loc_e = "N/A"
                if energies:
                    max_e_idx = max(energies, key=energies.get)
                    peak_energy = energies[max_e_idx]
                    loc_e = f"{max_e_idx[0]}_{j}_{max_e_idx[1]}"
                    
                metrics[comp][j]['bending'].append(bending_deg)
                metrics[comp][j]['twist'].append(twist_deg)
                metrics[comp][j]['energy'].append(peak_energy)
                
                # Store locations only if it's a new max (to retrieve easily later, simplified below)
                metrics[comp][j]['loc_b'].append(loc_b)
                metrics[comp][j]['loc_t'].append(loc_t)
                metrics[comp][j]['loc_e'].append(loc_e)

    log_and_print("-" * 65)
    log_and_print("Simulation completed.")
    
    # Calculate global peak G over the differentiated root trajectory
    # Double differentiation of Z position for absolute shock magnitude
    z_array = np.array(z_hist)
    if len(z_array) > 1:
        v_z = np.gradient(z_array, dt)
        a_z = np.gradient(v_z, dt)
        # The actual physics simulation tracks relative to earth freefall, so we pull max impact
        root_acc_history = np.abs(a_z) / 9.81
        max_g_force = float(np.max(root_acc_history))
    else:
        max_g_force = 0.0
    # Generate Terminal Summary Report and save to file
    # [Changed] log_and_print helper is now defined earlier and persistent.

    log_and_print("-" * 50)
    log_and_print(f"Peak Assembly G-Force: {max_g_force:.2f} G")
    log_and_print("-" * 50)

    for comp, j_data in metrics.items():
        # [NEW] 리지드 바디(Rigid Body) 및 단일 블록(Aux Mass) 제외 로직
        # 구조적 변형(굽힘, 비틂)은 최소 2개 이상의 블록이 존재해야 정의 가능합니다.
        # 블록이 1개뿐인 컴포넌트는 강체로 간주하고 리포트에서 제외합니다.
        if len(components[comp]) <= 1:
            continue

        log_and_print("=" * 83)
        log_and_print(f"[최대 구조 변형 지표 로컬라이징 리포트] - Body: {comp}")
        log_and_print("-" * 83)
        log_and_print(f"{'Row Index':<9} | {'Max Bending (deg) / Loc':<23} | {'Max Twist (deg) / Loc':<21} | {'Peak Energy (J) / Loc':<21}")
        log_and_print("-" * 83)
        
        j_keys = [k for k in j_data.keys() if isinstance(k, int)]
        for j in sorted(j_keys):
            hist = j_data[j]
            max_b = max(hist['bending'])
            idx_b = hist['bending'].index(max_b)
            loc_b = hist['loc_b'][idx_b]
            
            max_t = max(hist['twist'])
            idx_t = hist['twist'].index(max_t)
            loc_t = hist['loc_t'][idx_t]
            
            max_e = max(hist['energy'])
            idx_e = hist['energy'].index(max_e)
            loc_e = hist['loc_e'][idx_e]
            
            str_b = f"{max_b:>6.2f} (@ {loc_b:<7})"
            str_t = f"{max_t:>6.2f} (@ {loc_t:<7})"
            str_e = f"{max_e:>6.2f} (@ {loc_e:<7})"
            log_and_print(f"y = {j:<5} | {str_b:<23} | {str_t:<21} | {str_e:<21}")
        log_and_print("=" * 83)
        log_and_print("")
        
    # Plotting (only if requested)
    if config.get("plot_results", True):
        # [NEW] txt 파일 출력을 위한 공용 헬퍼 함수
        export_txt_path = os.path.join(output_dir, f'rds-{timestamp}_data.txt')
        f_txt = open(export_txt_path, "w", encoding="utf-8")
        
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as xlImage
            has_xl = True
            xl_wb = Workbook()
            xl_wb.remove(xl_wb.active)
            export_xl_path = os.path.join(output_dir, f'rds-{timestamp}_data.xlsx')
            xl_row_tracker = {}
        except ImportError:
            has_xl = False
            
        def export_plot_data(title, time_arr, data_dict, sheet_name=None, png_path=None):
            f_txt.write(f"#TITLE: {title}\n")
            legends = list(data_dict.keys())
            max_len = max([len(str(l)) for l in legends]) if legends else 0
            num_header_rows = max(1, (max_len + 18) // 19)
            
            for r in range(num_header_rows):
                header = f"#{'Time (s)':<18}" if r == 0 else f"#{'':<18}"
                for leg in legends:
                    header += f"{str(leg)[r*19:(r+1)*19]:<20}"
                f_txt.write(header + "\n")
            
            def fmt(v):
                s = f"{float(v):.8f}"
                if len(s) > 18:
                    s = f"{float(v):.8e}"
                return f"{s:<20}"
                
            for i in range(len(time_arr)):
                line_str = f" {fmt(time_arr[i]).strip():<19}"
                for leg in legends:
                    try: val = data_dict[leg][i]
                    except (IndexError, TypeError): val = 0.0
                    line_str += fmt(val)
                f_txt.write(line_str + "\n")
            f_txt.write("\n")
            
            if has_xl and sheet_name:
                safe_sheet = str(sheet_name)[:31]
                if safe_sheet in xl_wb.sheetnames:
                    ws = xl_wb[safe_sheet]
                else:
                    ws = xl_wb.create_sheet(title=safe_sheet)
                    if png_path and os.path.exists(png_path):
                        try:
                            img = xlImage(png_path)
                            ws.add_image(img, 'A1')
                        except: pass
                
                if safe_sheet not in xl_row_tracker:
                    xl_row_tracker[safe_sheet] = 11
                start_row = xl_row_tracker[safe_sheet]
                
                ws.cell(row=start_row, column=1, value=f"#TITLE: {title}")
                start_row += 1
                
                headers = ['Time (s)'] + legends
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=start_row, column=col_idx, value=str(h))
                    
                for i in range(len(time_arr)):
                    r_idx = start_row + 1 + i
                    ws.cell(row=r_idx, column=1, value=time_arr[i])
                    for col_idx, leg in enumerate(legends, 2):
                        try: val = float(data_dict[leg][i])
                        except: val = 0.0
                        ws.cell(row=r_idx, column=col_idx, value=val)
                xl_row_tracker[safe_sheet] = start_row + 1 + len(time_arr) + 1

        plt.figure(figsize=(10, 5))
        plt.plot(time_history, root_acc_history, label='Internal TV Accel (G)', color='black')
        plt.xlabel('Time (s)')
        plt.ylabel('G-Force')
        plt.title('Internal TV (OpenCell/Chassis) Peak Impact G-Force')
        plt.grid(True)
        plt.legend()
        plt.tight_layout()
        png1 = os.path.join(output_dir, 'rds-impact_gforce.png')
        plt.savefig(png1)
        plt.close()
        export_plot_data('Internal TV (OpenCell/Chassis) Peak Impact G-Force', time_history, {'Internal TV Accel (G)': root_acc_history}, 'Peak Impact G-Force', png1)
        
        # Ground Impact + Air Resistance combined plot (2 subplots)
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 8), sharex=True)
        fig.suptitle('Ground Impact Force & Air Resistance Forces')
        ax1.plot(time_history, ground_impact_hist, label='Ground Normal Force (N)', color='red')
        ax1.set_ylabel('Force (N)')
        ax1.set_title('Ground Impact')
        ax1.grid(True)
        ax1.legend()
        ax2.plot(time_history, air_drag_hist,    label=f'Drag  (Cd={config.get("air_cd_drag",1.05):.2f})', color='blue')
        ax2.plot(time_history, air_viscous_hist, label=f'Viscous (Cd_v={config.get("air_cd_viscous",0.0):.2f})', color='green')
        ax2.plot(time_history, air_squeeze_hist, label=f'Squeeze Film (k={config.get("air_coef_squeeze",1.0):.1f})', color='orange')
        ax2.set_xlabel('Time (s)')
        ax2.set_ylabel('Force (N)')
        ax2.set_title('Air Resistance Components')
        ax2.grid(True)
        ax2.legend()
        plt.tight_layout()
        png2 = os.path.join(output_dir, 'rds-ground_impact.png')
        plt.savefig(png2)
        plt.close()
        
        export_plot_data('Ground Impact Force', time_history, {'Ground Normal Force (N)': ground_impact_hist}, 'Ground Impact', png2)
        export_plot_data('Air Resistance Components', time_history, {
            f'Drag (Cd={config.get("air_cd_drag",1.05):.2f})': air_drag_hist,
            f'Viscous (Cd_v={config.get("air_cd_viscous",0.0):.2f})': air_viscous_hist,
            f'Squeeze Film (k={config.get("air_coef_squeeze",1.0):.1f})': air_squeeze_hist
        }, 'Ground Impact')
        
        # Motion All & Motion Z Setup
        pos_np = np.array(pos_hist) # (steps, 3)
        vel_np = np.array(vel_hist) # (steps, 6) -> [wx, wy, wz, vx, vy, vz]
        acc_np = np.array(acc_hist) # (steps, 6) -> [ax, ay, az, lin_ax, ...]
        
        c_pos_np = np.array(corner_pos_hist) # (steps, 8, 3)
        c_vel_np = np.array(corner_vel_hist) # (steps, 8, 3)
        c_acc_np = np.array(corner_acc_hist) # (steps, 8, 3)
        
        curves = {
            'Center': {'pos': pos_np, 'vel': vel_np[:, 3:6], 'acc': acc_np[:, 3:6]},
            'True_COG': {'pos': np.array(cog_pos_hist), 'vel': np.array(cog_vel_hist), 'acc': np.array(cog_acc_hist)},
            'CORNER_L-B-B': {'pos': c_pos_np[:,0,:], 'vel': c_vel_np[:,0,:], 'acc': c_acc_np[:,0,:]},
            'CORNER_L-B-F': {'pos': c_pos_np[:,1,:], 'vel': c_vel_np[:,1,:], 'acc': c_acc_np[:,1,:]},
            'CORNER_L-T-B': {'pos': c_pos_np[:,2,:], 'vel': c_vel_np[:,2,:], 'acc': c_acc_np[:,2,:]},
            'CORNER_L-T-F': {'pos': c_pos_np[:,3,:], 'vel': c_vel_np[:,3,:], 'acc': c_acc_np[:,3,:]},
            'CORNER_R-B-B': {'pos': c_pos_np[:,4,:], 'vel': c_vel_np[:,4,:], 'acc': c_acc_np[:,4,:]},
            'CORNER_R-B-F': {'pos': c_pos_np[:,5,:], 'vel': c_vel_np[:,5,:], 'acc': c_acc_np[:,5,:]},
            'CORNER_R-T-B': {'pos': c_pos_np[:,6,:], 'vel': c_vel_np[:,6,:], 'acc': c_acc_np[:,6,:]},
            'CORNER_R-T-F': {'pos': c_pos_np[:,7,:], 'vel': c_vel_np[:,7,:], 'acc': c_acc_np[:,7,:]}
        }
        ordered_keys = ['Center', 'True_COG', 'CORNER_L-T-F', 'CORNER_R-T-F', 'CORNER_L-T-B', 'CORNER_R-T-B', 'CORNER_L-B-F', 'CORNER_R-B-F', 'CORNER_L-B-B', 'CORNER_R-B-B']
        
        # rds-Motion_All.png
        fig, axs = plt.subplots(3, 3, figsize=(18, 12))
        fig.suptitle('Kinematics: Position, Velocity, Acceleration vs Time', fontsize=16)
        labels_row = ['Position (m)', 'Velocity (m/s)', 'Acceleration (m/s^2)']
        labels_col = ['X Axis', 'Y Axis', 'Z Axis']
        all_subplot_dicts = []
        for row in range(3):
            for col in range(3):
                ax = axs[row, col]
                subplot_dict = {}
                for k in ordered_keys:
                    metric = 'pos' if row == 0 else ('vel' if row == 1 else 'acc')
                    ax.plot(time_history, curves[k][metric][:, col], label=k)
                    subplot_dict[k] = curves[k][metric][:, col]
                if row == 2: ax.set_xlabel('Time (s)')
                if col == 0: ax.set_ylabel(labels_row[row])
                if row == 0: ax.set_title(labels_col[col])
                ax.grid(True)
                all_subplot_dicts.append((f"Kinematics - {labels_row[row]} - {labels_col[col]}", subplot_dict))
        axs[0, 2].legend(loc='upper left', bbox_to_anchor=(1.05, 1))
        plt.tight_layout(rect=[0, 0, 0.9, 1])
        png3 = os.path.join(output_dir, 'rds-Motion_All.png')
        plt.savefig(png3)
        plt.close()
        for i, (ttl, s_dict) in enumerate(all_subplot_dicts):
            export_plot_data(ttl, time_history, s_dict, 'Motion_All', png3 if i==0 else None)
        
        # rds-Motion_Z.png
        fig, axs = plt.subplots(1, 3, figsize=(18, 5))
        fig.suptitle('Kinematics (Z Axis Only): Position, Velocity, Acceleration vs Time', fontsize=16)
        titles_z = ['Z Position (m)', 'Z Velocity (m/s)', 'Z Acceleration (m/s^2)']
        z_subplot_dicts = []
        for i, metric in enumerate(['pos', 'vel', 'acc']):
            ax = axs[i]
            subplot_dict = {}
            for k in ordered_keys:
                ax.plot(time_history, curves[k][metric][:, 2], label=k)
                subplot_dict[k] = curves[k][metric][:, 2]
            ax.set_xlabel('Time (s)')
            ax.set_ylabel(titles_z[i])
            ax.grid(True)
            z_subplot_dicts.append((f"Kinematics (Z Axis Only) - {titles_z[i]}", subplot_dict))
        axs[2].legend(loc='upper left', bbox_to_anchor=(1.05, 1))
        plt.tight_layout(rect=[0, 0, 0.9, 1])
        png4 = os.path.join(output_dir, 'rds-Motion_Z.png')
        plt.savefig(png4)
        plt.close()
        for i, (ttl, s_dict) in enumerate(z_subplot_dicts):
            export_plot_data(ttl, time_history, s_dict, 'Motion_Z', png4 if i==0 else None)

        # Plot metric per component max over time
        for comp, j_data in metrics.items():
            # [NEW] 블록이 1개인 강체는 시각화 리포트에서도 제외
            if len(components[comp]) <= 1:
                continue

            j_keys = [k for k in j_data.keys() if isinstance(k, int)]
            if len(j_keys) > 0:
                plt.figure(figsize=(10, 5))
                
                # Find maximums across all J rows dynamically
                all_b_hist = np.array([j_data[j]['bending'] for j in j_keys])
                all_t_hist = np.array([j_data[j]['twist'] for j in j_keys])
                
                max_b_time = np.max(all_b_hist, axis=0)
                max_t_time = np.max(all_t_hist, axis=0)
                
                # Find the location string of the absolute global peak to display in legend
                max_b_idx = np.argmax(max_b_time)
                row_idx_b = np.argmax(all_b_hist[:, max_b_idx])
                global_loc_b = j_data[j_keys[row_idx_b]]['loc_b'][max_b_idx]
                
                max_t_idx = np.argmax(max_t_time)
                row_idx_t = np.argmax(all_t_hist[:, max_t_idx])
                global_loc_t = j_data[j_keys[row_idx_t]]['loc_t'][max_t_idx]
                
                plt.plot(time_history, max_b_time, label=f'Max Bending (deg) [{global_loc_b}]')
                plt.plot(time_history, max_t_time, label=f'Max Twisting (deg) [{global_loc_t}]')
                plt.xlabel('Time (s)')
                plt.ylabel('Angle (deg)')
                plt.title(f'{comp} Structural Deformation (Max)')
                plt.grid(True)
                plt.legend()
                plt.tight_layout()
                png_def = os.path.join(output_dir, f'rds-{comp}_deformation.png')
                plt.savefig(png_def)
                plt.close()
                export_plot_data(f'{comp} Structural Deformation (Max)', time_history, {
                    f'Max Bending [{global_loc_b}]': max_b_time,
                    f'Max Twisting [{global_loc_t}]': max_t_time
                }, f'{comp[:15]}_deformation', png_def)
            
            # New: All Blocks Angle Plot
            all_blocks = j_data.get('all_blocks_angle', {})
            block_noms = j_data.get('block_nominals', {})
            
            # [NEW] 블록이 1개뿐인 경우 상대적인 각도 변화 차트가 무의미하므로 제외
            if len(all_blocks) > 0 and len(components[comp]) > 1:
                fig = plt.figure(figsize=(14, 6))
                
                # Create main plot for angle curves
                ax_main = fig.add_axes([0.05, 0.1, 0.65, 0.8])
                num_blocks = len(all_blocks)
                cols_legend = min(4, max(1, num_blocks // 10))
                
                block_dict = {}
                for block_idx, a_hist in all_blocks.items():
                    legend_name = f"{block_idx[0]}-{block_idx[1]}-{block_idx[2]}"
                    ax_main.plot(time_history, a_hist, label=legend_name, linewidth=1.0)
                    block_dict[legend_name] = a_hist
                ax_main.set_xlabel('Time (s)')
                ax_main.set_ylabel('Def. Angle (deg)')
                ax_main.set_title(f'{comp} Block Angle Deformation (Relative to Root)')
                ax_main.grid(True)
                
                # Put Legend outside
                ax_main.legend(loc='upper left', bbox_to_anchor=(1.0, 1.0), ncol=cols_legend, fontsize=6)
                
                # Inset 3D scatter plot for index guide
                # Move slightly further down-right to avoid any potential clash
                ax_inset = fig.add_axes([0.72, 0.05, 0.25, 0.25], projection='3d')
                xs, ys, zs = [], [], []
                labels = []
                for b_idx, nom in block_noms.items():
                    xs.append(nom[0])
                    ys.append(nom[1])
                    zs.append(nom[2])
                    labels.append(f"{b_idx[0]}-{b_idx[1]}-{b_idx[2]}")
                ax_inset.scatter(xs, ys, zs, color='blue', alpha=0.5, s=10)
                
                # Subsample labels if too many to avoid clutter
                step = max(1, len(labels) // 20)
                for i in range(0, len(labels), step):
                    ax_inset.text(xs[i], ys[i], zs[i], labels[i], size=5, zorder=1, color='k')
                
                ax_inset.set_title("i-j-k Reference")
                ax_inset.set_xticks([])
                ax_inset.set_yticks([])
                ax_inset.set_zticks([])
                
                png_def_all = os.path.join(output_dir, f'rds-{comp}_deformation_all.png')
                plt.savefig(png_def_all)
                plt.close()
                export_plot_data(f'{comp} Block Angle Deformation (Relative to Root)', time_history, block_dict, f'{comp[:15]}_def_all', png_def_all)
            
        f_txt.close()
        if has_xl:
            xl_wb.save(export_xl_path)
            log_and_print(f"  >> Excel saved to: {os.path.basename(export_xl_path)}")
        
    # Copy the XML file used for simulation to the output directory
    if xml_path and os.path.exists(xml_path):
        shutil.copy(xml_path, os.path.join(output_dir, os.path.basename(xml_path)))
        log_and_print(f"  >> Results saved to: {os.path.basename(output_dir)}")
    
    # [Important] Callback 해제 (메모리 릭 및 충돌 방지)
    mujoco.set_mjcb_control(None)
    
    sim_result = DropSimResult(
        config=config,
        metrics=metrics,
        max_g_force=max_g_force,
        time_history=time_history,
        z_hist=z_hist,
        root_acc_history=root_acc_history,
        corner_acc_hist=corner_acc_hist,
        pos_hist=pos_hist,
        vel_hist=vel_hist,
        acc_hist=acc_hist,
        cog_pos_hist=cog_pos_hist,
        cog_vel_hist=cog_vel_hist,
        cog_acc_hist=cog_acc_hist,
        corner_pos_hist=corner_pos_hist,
        corner_vel_hist=corner_vel_hist,
        ground_impact_hist=ground_impact_hist,
        air_drag_hist=air_drag_hist,
        air_viscous_hist=air_viscous_hist,
        air_squeeze_hist=air_squeeze_hist
    )
    result_path = os.path.join(output_dir, f'rds-{timestamp}_result.pkl')
    sim_result.save(result_path)
    log_and_print(f"  >> Python 데이터 오브젝트(DropSimResult) 저장 완료: {os.path.basename(result_path)}")
    
    return sim_result, time_history, z_hist, vel_hist, root_acc_history, corner_acc_hist, max_g_force, metrics

def calculate_required_aux_masses(cfg, target_mass, target_cog, target_moi):
    """
    현재 모델의 관성 특성을 분석하고, 목표로 하는 질량/CoG/MoI를 달성하기 위해 
    다수의 보정 질량(aux_masses) 리스트를 계산하여 반환합니다.
    단일 블록으로는 달성하기 어려운 복합 관성을 맞추기 위해 8개의 대칭된
    포인트 질량체(Point Masses)로 목표치를 완벽히 분배합니다.
    """
    from run_discrete_builder import create_model
    import math
    
    # 1. 기저 모델 관성 획득 (보정 질량 제외)
    temp_cfg = cfg.copy()
    temp_cfg["chassis_aux_masses"] = []  
    
    # create_model 호출하여 현재 상태 측정
    _, m_base, c_base, i_base, _ = create_model("temp_calc.xml", config=temp_cfg, logger=lambda x: None)
    
    m_base = float(m_base)
    c_base = np.array(c_base)
    i_base = np.array(i_base) # [Ixx, Iyy, Izz]

    # 2. 목표값 처리 (None일 경우 Baseline 값 사용)
    target_mass = target_mass if target_mass is not None else m_base
    target_cog  = target_cog  if target_cog  is not None else c_base
    target_moi  = target_moi  if target_moi  is not None else i_base

    # 필요한 총 보정 질량
    m_aux = target_mass - m_base
    
    # 보정할 분량이 거의 없는 경우 빈 리스트 반환
    if abs(m_aux) < 1e-6:
        # 질량은 맞는데 CoG나 MoI가 다를 수도 있으므로 체크
        # 하지만 m_aux=0이면 점질량으로 보정 불가. 최소 질량 요구.
        if (np.allclose(target_cog, c_base, atol=1e-5) and 
            np.allclose(target_moi, i_base, atol=1e-5)):
            return []
        else:
            raise ValueError(f"질량 변화 없이 CoG/MoI만 보정하는 것은 물리적으로 불가능합니다 (최소한의 추가 질량이 필요함).")

    if m_aux < 0:
        raise ValueError(f"목표 질량({target_mass})이 현재 질량({m_base})보다 작습니다. (감량 보정은 지원하지 않음)")

    # 3. 추가 질량체 군집의 목표 CoG 연산
    pos_aux = (np.array(target_cog) * target_mass - m_base * c_base) / m_aux

    # 4. 관성 모멘트 평행축 정리 역산
    def shift_moi(m, i_cg, cg, target_p):
        d = target_p - cg
        ix = i_cg[0] + m * (d[1]**2 + d[2]**2)
        iy = i_cg[1] + m * (d[0]**2 + d[2]**2)
        iz = i_cg[2] + m * (d[0]**2 + d[1]**2)
        return np.array([ix, iy, iz])

    i_base_at_target = shift_moi(m_base, i_base, c_base, target_cog)
    i_aux_req_at_target = np.array(target_moi) - i_base_at_target
    
    # pos_aux 기준에서 가져야할 순수 Inertia 
    i_aux_req_at_own_cog = i_aux_req_at_target - shift_moi(m_aux, [0,0,0], pos_aux, target_cog)
    
    # 이론상 불가능할 경우(음수) 최소 오차를 위해 0으로 제한 (물리적 한계 보정)
    i_res = np.maximum(i_aux_req_at_own_cog, 1e-9)

    # 5. 8개의 질량체로 대칭 분배
    m_each = m_aux / 8.0
    
    # 기하학적 형상(단일 박스)의 구속에 얽매이지 않고, 대칭된 8개의 점으로 텐서 완벽 분해
    # 8-Point Mass 위치 오프셋 해석해
    dx_sq = (i_res[1] + i_res[2] - i_res[0]) / (2.0 * m_aux)
    dy_sq = (i_res[0] + i_res[2] - i_res[1]) / (2.0 * m_aux)
    dz_sq = (i_res[0] + i_res[1] - i_res[2]) / (2.0 * m_aux)
    
    dx = math.sqrt(max(0, dx_sq))
    dy = math.sqrt(max(0, dy_sq))
    dz = math.sqrt(max(0, dz_sq))
    
    s = [0.01, 0.01, 0.01] # 충돌이나 시각을 방해하지 않는 매우 작은 형상 더미
    
    aux_masses = []
    idx = 1
    for sign_x in [-1, 1]:
        for sign_y in [-1, 1]:
            for sign_z in [-1, 1]:
                px = pos_aux[0] + sign_x * dx
                py = pos_aux[1] + sign_y * dy
                pz = pos_aux[2] + sign_z * dz
                
                aux_masses.append({
                    "name": f"AutoB_{idx}",
                    "pos": [float(round(px, 5)), float(round(py, 5)), float(round(pz, 5))],
                    "mass": float(round(m_each, 5)),
                    "size": s
                })
                idx += 1

    return aux_masses

def test_run_case_1():
    # Get basic defaults, override specifically if needed    
    cfg = get_default_config()
    # 예: 전면 하단 꼭짓점 낙하 자세로 설정
    cfg["drop_mode"] = "L-F-B" 
    cfg["include_paperbox"] = False # 로컬 테스트용 오버라이드
    cfg["drop_height"] = 0.5    
    cfg["plot_results"] = True    
    cfg["only_generate_xml"] = False # [NEW] True 설정 시 시뮬레이션을 돌리지 않고 XML 모델 생성 및 저장만 수행

    # [COMPONENTS OPTIONS] 부품별 해상도(div) 및 결합 방식(use_weld) 설정
    # use_weld=False 설정 시, 해당 부품은 내부 구속조건이 없는 '단일 강체'로 취급되어 연산 속도가 비약적으로 향상됩니다.
    cfg["chassis_div"]      = [4, 4, 1]    # 샤시 분할 수
    cfg["chassis_use_weld"] = False        # 샤시는 강체로 취급 (속도 향상)
    
    cfg["oc_div"]           = [4, 4, 1]    # 오픈셀(패널) 분할 수
    cfg["oc_use_weld"]      = False         # 오픈셀은 강체로 취급 (속도 향상)
    
    cfg["occ_div"]          = [4, 4, 1]    # 테이프(Cohesive) 분할 수
    cfg["occ_use_weld"]     = False         # 테이프는 유연성 유지를 위해 Weld 사용
    
    cfg["cush_div"]         = [5, 4, 3]    # 쿠션 분할 수
    cfg["cush_use_weld"]    = True         # 쿠션은 변형이 중요하므로 Weld 사용
    
    cfg["box_div"]          = [5, 4, 2]    # 외곽 박스 분할 수 (성능을 위해 낮춤)
    cfg["box_use_weld"]     = False        # 박스는 강체로 취급
    
    cfg["tape_solref"] = "0.05 1.0"
    
    # [NEW] 영구 변형 테스트 활성화
    cfg["enable_plasticity"] = True
    cfg["plasticity_ratio"] = 0.5

    cfg["sim_duration"] = 1.0

    cfg["mass_paper"] = 4.0
    cfg["mass_cushion"] = 2.0
    cfg["mass_oc"] = 5.0
    cfg["mass_occ"] = 0.1
    cfg["mass_chassis"] = 10.0

    cfg["ground_friction"] = 0.01 # 바닥 마찰계수 기본값 0.2
    
    # [RECOMMENDED] 딱딱한 바닥과의 충돌 시 관통 방지를 위한 설정
    # ground_solref: [timeconst, dampratio] -> timeconst가 작을수록 딱딱함. (0.002 미만은 비권장)
    cfg["ground_solref_stiff"] = 0.005  # 0.01에서 0.004로 강화 (더 딱딱한 바닥)
    cfg["ground_solref_damp"]  = 0.01    # Critical Damping
    
    # ground_solimp: [dmin, dmax, width, midpoint, power] 
    # dmax를 0.999로 높여 최대 압축 시 반발력을 극대화하고, width를 줄여 즉각 반응하게 함
    cfg["ground_solimp"] = "0.9 0.999 0.001 0.5 2"
    
    # [SIMULATION OPTIONS] MuJoCo 물리 엔진 및 솔버 관련 상세 설정 (sim_*)
    cfg["sim_integrator"] = "Euler" # 통합기 (Euler, RK4, implicit, implicitfast 등)
    cfg['sim_duration'] = 3.0
    cfg["sim_timestep"]   = 0.0013         # 시뮬레이션 타임스텝 (s)
    cfg["sim_iterations"] = 50             # 솔버 최대 반복 횟수 (1~200)
    cfg["sim_noslip_iterations"] = 0       # 마찰 고정(미끄러짐 방지)을 위한 추가 반복 횟수
    cfg["sim_tolerance"]  = 1e-5           # 솔버 수렴 오차 허용치 (속도를 위해 약간 완화)
    cfg["sim_gravity"]    = [0, 0, -9.81]  # 중력 가속도 [x, y, z]
    cfg["sim_nthread"]    = 4              # [NEW] 멀티코어 사용 (코어 수에 맞춰 조절)
    cfg["sim_impratio"]   = 5.0             # 기본 1.0. 값을 5~10 정도로 크게 주면 관통을 강력하게 봉쇄합니다.

    # [SOLVER OPTIONS] MuJoCo 물리 엔진 및 솔버 관련 상세 설정 (sol_*)
    cfg["cush_solref_stiff"] = 0.07 # 쿠션 timeconst
    cfg["cush_solref_damp"]  = 0.7 # 쿠션 dampratio

    # air fluidic force
    cfg["air_density"]      = 1.225     # 공기 밀도 (kg/m^3, 20도 1atm)
    cfg["air_viscosity"]    = 1.81e-5   # 공기 동점성계수 (Pa.s)
    cfg["air_cd_drag"]      = 1.05      # Blunt drag 계수 (박스 형태 기준 1.0~1.2)
    cfg["air_cd_viscous"]   = 0.01      # Slender(점성) drag 계수 (박스는 보통 0)
    cfg["air_coef_squeeze"] = 0.2       # Squeeze Film 효과 강도 배율 (0=비활성화)
    cfg["air_squeeze_hmax"] = 0.20      # Squeeze Film 활성화 최대 높이 (m)
    cfg["air_squeeze_hmin"] = 0.005      # Squeeze Film 최소 높이 (분모 안전값, m)
    cfg["enable_air_drag"]    = True   # MuJoCo 빌트인 Drag/Viscous 활성화 여부
    cfg["enable_air_squeeze"] = True   # 수동 Squeeze Film 활성화 여부

    # [STEP 1] 보정 전 설계 원안 검토 (Baseline Review)
    print_inertia_report(cfg, title="Baseline Inertia Report (Pre-Correction)", logger=print)
    
    # [STEP 2] 자동 질량 보정 적용 (Target 맞춤)
    target_mass = 25.0
    target_cog = [0.0, -0.01, -0.01]
    #target_moi = [3.0, 8.0, 10.0]  # [Ixx, Iyy, Izz]
    target_moi = None

    try:
        # 예시: CoG만 보정하고 싶다면 target_mass=None, target_moi=None 입력 가능
        aux_items = calculate_required_aux_masses(cfg, target_mass, target_cog, target_moi)
        cfg["chassis_aux_masses"] = aux_items
        print(f"\n>> [Inertia Balancer] Target 기반 보정 질량 (8-Point 다중 질량체) 생성 완료:")
        print(f"      - 생성된 개체 수: {len(aux_items)}EA, 총 추가 질량: {sum(x['mass'] for x in aux_items):.4f}kg")
        if aux_items:
            for aux in aux_items[:2]:
                print(f"      - 예시: {aux['name']} / Pos: {aux['pos']} / Mass: {aux['mass']}kg")
            print("      ... 등격(Symmetric) 배치됨")
        else:
            print("      - 보정이 필요하지 않은 상태입니다 (Baseline 유지).")
    except Exception as e:
        print(f"\n>> [Inertia Balancer] 보정 계산 실패 (기본값 사용): {e}")
        cfg["chassis_aux_masses"] = []
    
    # [STEP 3] 실행 전 최종 확인 (선택 사항)
    print("\n>> 보정이 적용된 최종 모델 구성을 확인합니다...")
    print_inertia_report(cfg, title="Final Balanced Inertia Report", logger=print)

    # [STEP 4] 실행 시 뷰어 사용 여부 확인 후 시뮬레이션 시작
    cfg["use_viewer"] = ask_use_viewer()
    run_simulation(cfg)


if __name__ == "__main__":
    test_run_case_1()

'''
1. 재료별 추천 
solref 값 가이드
- 재료 유형	추천 timeconst (강성)	추천 dampratio (감쇠)	특징 및 설명
- Rigid (금속, 딱딱한 플라스틱)	0.002 ~ 0.005	1.0 ~ 2.0	변형이 거의 없는 소자. 타임스텝의 2배(0.002)가 수치적 한계치입니다.
- Semi-Rigid (골판지, 강화 플라스틱)	0.01 ~ 0.02	1.0	약간의 탄성이 느껴지지만 기본적으로 형태를 유지해야 하는 경우입니다.
- Adhesive/Tape (점착제, 테이프)	0.005 ~ 0.02	1.0 ~ 1.5	현재 프로젝트의 Cohesive 층에 해당합니다. 결합력이 강할수록 낮은 값을 씁니다.
- Foam/Cushion (스티로폼, EPP)	0.02 ~ 0.1	0.5 ~ 1.0	충격을 흡수하며 눌려야 하는 재질입니다. 값이 클수록 말랑해집니다.
- Soft Rubber (부드러운 고무)	0.05 ~ 0.2	0.1 ~ 0.5	복원력이 강하지만 아주 부드럽게 눌리는 재질입니다.

2. 파라미터별 세부 조정 팁
Timeconst (첫 번째 값)
- 의미: 스프링-댐퍼 시스템이 평형 상태로 돌아오는 데 걸리는 시간 상수입니다.
- Rule of Thumb: 시뮬레이션 타임스텝(dt)의 최소 2배 이상을 권장합니다.
- dt=0.001일 때 0.002가 물리적 강성의 최대치입니다. 이보다 낮으면 수치적 폭발(Explosion)이 일어날 확률이 급격히 높아집니다.
- 조정: 물체가 너무 잘 뚫고 지나가면(Penetration) 이 값을 낮추고, 시뮬레이션이 불안정하게 떨리면 이 값을 높여야 합니다.

Dampratio (두 번째 값)
- 의미: 감쇠비입니다. 1.0이 임계 감쇠(Critical Damping)입니다.
- 1.0: 에너지를 가장 안정적으로 소산시키며 진동 없이 멈춥니다. (대부분의 조립체 권장)
- > 1.0 (Over-damping): 충격 시 튕겨나가는 현상을 억제하고 싶을 때 사용합니다. 낙하 시험에서 바닥과의 접촉 등에 유리합니다.
- < 1.0 (Under-damping): 물체가 통통 튀는 탄성 효과를 주고 싶을 때 사용합니다.

3. 현재 프로젝트(WHToolsBox) 적용 예시
사용자께서 방금 수정하신 값들을 기준으로 분석해 보면 다음과 같습니다.

- Chassis / TV Panel (0.01): 상당히 단단한 구조체로 적절한 설정입니다.
- Tape (0.05): 이전(0.005)보다 10배 더 말랑하게 수정하셨습니다. 이 경우 285G 같은 고충격에서는 테이프가 고무줄처럼 늘어나면서 부품 분리가 일어날 가능성이 매우 높습니다. (분리 문제가 다시 발생한다면 0.01 이하로 낮추는 것을 권장합니다.)
- Cushion (0.1): 아주 부드러운 완충재 설정입니다. 충격 흡수량은 많아지지만 패널이 쿠션을 깊게 파고들 수 있습니다.


1. i-j-k reference 그림과 legend가 겹치는 문제가 있는데 해결할 수 있을까?
2. 결과를 저장하는 python 변수들이 있다면, 이를 저장해놓았다가 다음에 읽어서 처리할 수 있으면 좋겠다. 이를 관장할 수 있는 데이터 클래스를 만들어 파일 저장, 읽기 등과 함게 데이터 처리, 가공, txt 및 excel 저장 등도 총괄 할 수 있도록 하자.

그리고 시뮬레이션 후에 rds 폴더에 해당 결과도 저장해놓고, 다음에 읽어서 데이터를 다뤌 수 있도록 해달라. 

우리는 나중에 주어진 시험 결과(예를 들면, 상자 모서리의 시간에 대한 좌표 변화)와 매칭되는 시뮬레이션 결과를 얻기 위해서 파라미터들을 자동으로 찾아 조정하는 기능을 구현할 게획이다. 이 계획도 참고하라.

'''

