import mujoco
import mujoco.viewer
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import os
import json
import time
import shutil
from datetime import datetime
from run_discrete_builder import create_model, get_default_config
import pickle
from dataclasses import dataclass, field
from typing import Any, Dict, List
import math
import io

@dataclass
class DropSimResult:
    """시뮬레이션 전체 결과 데이터를 담는 통합 클래스 (최적화 모델 매칭용)"""
    config: Dict[str, Any]
    metrics: Dict[str, Any]
    max_g_force: float
    time_history: List[float]
    z_hist: List[float]
    root_acc_history: List[Any]
    corner_acc_hist: List[Any]
    
    pos_hist: List[Any] = field(default_factory=list)
    vel_hist: List[Any] = field(default_factory=list)
    acc_hist: List[Any] = field(default_factory=list)
    
    cog_pos_hist: List[Any] = field(default_factory=list)
    cog_vel_hist: List[Any] = field(default_factory=list)
    cog_acc_hist: List[Any] = field(default_factory=list)
    
    corner_pos_hist: List[Any] = field(default_factory=list)
    corner_vel_hist: List[Any] = field(default_factory=list)
    
    ground_impact_hist: List[float] = field(default_factory=list)
    air_drag_hist: List[float] = field(default_factory=list)
    air_viscous_hist: List[float] = field(default_factory=list)
    air_squeeze_hist: List[float] = field(default_factory=list)
    
    def save(self, filepath: str):
        with open(filepath, "wb") as f:
            pickle.dump(self, f)
            
    @classmethod
    def load(cls, filepath: str):
        with open(filepath, "rb") as f:
            return pickle.load(f)

# =====================================================================
# [NEW] 전역 로깅 설정 (모든 터미널 출력을 리포트 파일에 저장)
# =====================================================================
_report_file_path = None

def log_and_print(msg):
    """
    터미널에 출력함과 동시에, 현재 진행 중인 시뮬레이션의 
    summary_report.txt 파일이 설정되어 있다면 해당 파일에도 기록합니다.
    """
    content = str(msg)
    print(content) # 실제 터미널 출력
    
    global _report_file_path
    if _report_file_path:
        try:
            # 'a' (append) 모드로 열어 파일 끝에 추가 기록
            with open(_report_file_path, "a", encoding="utf-8") as rf:
                rf.write(content + "\n")
        except Exception as e:
            # 로깅 자체의 에러는 터미널에 최소한으로 출력
            print(f"!!! [Logging Failed] {e}")

# =====================================================================
# [NEW] 리포팅 및 설정을 위한 유틸리티 함수
# =====================================================================
def format_config_report(config, timestamp):
    """설정(Config) 데이터를 보기 좋게 문자열로 정리하여 반환합니다."""
    lines = []
    lines.append("\n" + "="*90)
    lines.append(f"  [MuJoCo Drop Simulation - Configuration Summary]  -  {timestamp}")
    lines.append("="*90)
    
    # 주요 카테고리별 정렬 정의
    categories = {
        "1. 낙하 및 실행 조건 (Drop & Run)": ["drop_mode", "drop_direction", "drop_height", "sim_duration", "only_generate_xml", "plot_results", "use_viewer"],
        "2. 구조물 및 조립 (Assembly)":       ["include_paperbox", "include_cushion", "box_use_weld", "cush_use_weld", "oc_use_weld", "occ_use_weld", "chassis_use_weld"],
        "3. 오픈셀/테이프 형상 (OC/Tape Geo)": ["oc_div", "oc_d", "occ_div", "occ_d", "occ_ithick"],
        "4. 공통 형상 파라미터 (Common Geo)": ["box_w", "box_h", "box_d", "box_thick", "box_div", "cush_div", "cush_gap", "assy_div", "assy_w", "assy_h", "chassis_div", "chas_d"],
        "5. 질량 및 물성 (Material)":        ["mass_paper", "mass_cushion", "mass_oc", "mass_occ", "mass_chassis", "ground_friction", "enable_plasticity", "plasticity_ratio", "cush_yield_stress",
                                              "ground_solref", "ground_solref_timec", "ground_solref_damprr", "ground_solimp", 
                                              "cush_solref", "cush_solref_timec", "cush_solref_damprr",
                                              "cush_weld_solref", "cush_weld_solref_timec", "cush_weld_solref_damprr", "cush_weld_solimp",
                                              "cush_contact_solref", "cush_contact_solimp", "cush_corner_solref", "cush_corner_solimp", 
                                              "tape_weld_solref", "tape_solref", "cell_weld_solref", "cell_solref"],
        "6. 솔버 및 물리 파라미터 (Physics)": ["sim_integrator", "sim_timestep", "sim_iterations", "sim_impratio", "sim_nthread", "sim_gravity", "sim_tolerance", "sim_noslip_iterations"],
        "7. 공기 역학 (Aerodynamics)":       ["enable_air_drag", "enable_air_squeeze", "air_density", "air_viscosity", "air_cd_drag", "air_cd_viscous", "air_coef_squeeze", "air_squeeze_hmax", "air_squeeze_hmin"]
    }
    
    # Mapping of parameter names to English descriptions
    descriptions = {
        "drop_mode": "Standard drop test mode (PARCEL, LTL, CUSTOM, etc.)",    # 표준 낙하 시험 모드 (PARCEL, LTL, CUSTOM 등)
        "drop_direction": "Specific drop orientation (e.g., front-bottom-left)", # 구체적인 낙하 방향 (예: front-bottom-left)
        "drop_height": "Drop height from ground (m)",                         # 지면으로부터의 낙하 높이 (m)
        "sim_duration": "Total simulation duration (s)",                      # 총 시뮬레이션 시간 (s)
        "only_generate_xml": "Generate XML and skip simulation",              # XML 모델만 생성하고 시뮬레이션은 건너뜀
        "plot_results": "Enable plotting of simulation results",              # 시뮬레이션 결과 그래프 출력 여부
        "use_viewer": "Enable MuJoCo passive viewer GUI",                    # MuJoCo 패시브 뷰어 GUI 활성화
        "include_paperbox": "Include outer packaging box",                    # 외곽 포장 박스 포함 여부
        "include_cushion": "Include cushion components",                      # 완충재(쿠션) 포함 여부
        "box_use_weld": "Box as multiple bodies+weld (True) or single rigid (False)", # 박스 강체화 여부 (True: 다수 바디+Weld, False: 단일 강체)
        "cush_use_weld": "Cushion as multiple bodies+weld (True) or single rigid (False)", # 쿠션 강체화 여부 (True: 다수 바디+Weld, False: 단일 강체)
        "oc_use_weld": "Use internal welds for OpenCell",                    # 오픈셀 내부 Weld 사용 여부
        "occ_use_weld": "Use internal welds for Cohesive layer",              # 접착층 내부 Weld 사용 여부
        "chas_use_weld": "Use internal welds for Chassis",                    # 샤시 내부 Weld 사용 여부
        "chassis_use_weld": "Use internal welds for Chassis",                 # 샤시 내부 Weld 사용 여부 (중복 키)
        "box_w": "Packaging box width (m)",                                   # 포장 박스 가로 길이 (m)
        "box_h": "Packaging box height (m)",                                  # 포장 박스 세로 길이 (m)
        "box_d": "Packaging box depth (m)",                                   # 포장 박스 깊이 (m)
        "box_thick": "Packaging box wall thickness (m)",                      # 포장 박스 벽 두께 (m)
        "box_div": "Box discretization resolution [X, Y, Z]",                 # 박스 분할 해상도 [X, Y, Z]
        "cush_div": "Cushion discretization resolution [X, Y, Z]",            # 쿠션 분할 해상도 [X, Y, Z]
        "cush_gap": "Clearance gap between components (m)",                   # 부품 간 간격 (m)
        "assy_div": "Assembly discretization resolution [X, Y, Z]",           # 조립체 분할 해상도 [X, Y, Z]
        "assy_w": "Core assembly width (m)",                                  # 조립체 가로 치수 (m)
        "assy_h": "Core assembly height (m)",                                 # 조립체 세로 치수 (m)
        "oc_div": "OpenCell discretization resolution [X, Y, Z]",             # 오픈셀 분할 해상도 [X, Y, Z]
        "oc_d": "OpenCell thickness (m)",                                     # 오픈셀 두께 (m)
        "occ_div": "Cohesive layer discretization resolution [X, Y, Z]",       # 접착층 분할 해상도 [X, Y, Z]
        "occ_d": "Cohesive layer thickness (m)",                              # 접착층 두께 (m)
        "occ_ithick": "Cohesive layer initial thickness (m)",                 # 접착층 초기 두께 (m)
        "chassis_div": "Chassis discretization resolution [X, Y, Z]",          # 샤시 분할 해상도 [X, Y, Z]
        "chas_d": "Chassis thickness (m)",                                    # 샤시 두께 (m)
        "mass_paper": "Packaging box mass (kg)",                             # 포장 박스 질량 (kg)
        "mass_cushion": "Cushion mass (kg)",                                 # 쿠션 질량 (kg)
        "mass_oc": "OpenCell panel mass (kg)",                                # 오픈셀 패널 질량 (kg)
        "mass_occ": "Cohesive/Tape layer mass (kg)",                         # 접착/테이프층 질량 (kg)
        "mass_chassis": "TV Chassis mass (kg)",                              # TV 샤시 질량 (kg)
        "ground_friction": "Ground friction coefficient",                    # 바닥 마찰 계수
        "enable_plasticity": "Enable permanent plastic deformation",           # 소성 변형(영구 변형) 활성화
        "plasticity_ratio": "Plastic deformation ratio (0~1)",                # 소성 변형 비율 (0~1)
        "cush_yield_stress": "Yield stress threshold for cushion plasticity (MPa)", # 쿠션 소성 변형 임계값 (MPa)
        "cush_yield_strain": "Yield strain threshold for cushion plasticity", # 쿠션 소성 변형 임계 strain
        "sim_integrator": "Numerical integrator type (Euler, RK4, etc.)",    # 수치 적분기 방식 (Euler, RK4 등)
        "sim_timestep": "Simulation time step (s)",                           # 시뮬레이션 시간 간격 (s)
        "sim_iterations": "Max solver iterations per step",                  # 스텝당 최대 솔버 반복 횟수
        "sim_impratio": "Penetration constraint ratio (5~10 for rigid)",      # 관통 제약 배율 (강체는 5~10 권장)
        "sim_nthread": "Number of CPU threads for computation",               # 연산용 CPU 스레드 수
        "sim_gravity": "Gravity acceleration vector [x, y, z]",               # 중력 가속도 벡터 [x, y, z]
        "sim_tolerance": "Solver convergence tolerance",                      # 솔버 수렴 오차 허용치
        "sim_noslip_iterations": "Iterations for friction constraint",         # 마찰 제약(No-slip) 반복 횟수
        "ground_solref": "Ground contact solver (timeconst, dampratio)",      # 바닥 접촉 솔버 (타임상수, 감쇠비)
        "ground_solref_timec": "Ground contact time constant param",          # 바닥 접촉 타임 상수 파라미터
        "ground_solref_damprr": "Ground contact damping ratio param",          # 바닥 접촉 감쇠비 파라미터
        "ground_solimp": "Ground contact solver impedance (dmin, dmax, ...)", # 바닥 접촉 솔버 임피던스
        "cush_solref": "Cushion default solver reference",                    # 쿠션 기본 솔버 레퍼런스
        "cush_solref_timec": "Cushion time constant param",                   # 쿠션 타임 상수 파라미터
        "cush_solref_damprr": "Cushion damping ratio param",                   # 쿠션 감쇠비 파라미터
        "cush_weld_solref": "Cushion internal weld solver reference",         # 쿠션 내부 Weld 솔버 레퍼런스
        "cush_weld_solref_timec": "Cushion weld time constant param",         # 쿠션 Weld 타임 상수 파라미터
        "cush_weld_solref_damprr": "Cushion weld damping ratio param",         # 쿠션 Weld 감쇠비 파라미터
        "cush_weld_solimp": "Cushion internal weld solver impedance",         # 쿠션 내부 Weld 솔버 임피던스
        "cush_contact_solref": "Cushion center contact solver reference",      # 쿠션 중앙부 접촉 솔버 레퍼런스
        "cush_contact_solimp": "Cushion center contact solver impedance",      # 쿠션 중앙부 접촉 솔버 임피던스
        "cush_corner_solref": "Cushion edge/corner contact solver reference",   # 쿠션 모서리 접촉 솔버 레퍼런스
        "cush_corner_solimp": "Cushion edge/corner contact solver impedance",   # 쿠션 모서리 접촉 솔버 임피던스
        "tape_weld_solref": "Tape internal weld solver reference",            # 테이프 내부 Weld 솔버 레퍼런스
        "tape_solref": "Tape general contact solver reference",               # 테이프 일반 접촉 솔버 레퍼런스
        "cell_weld_solref": "OpenCell internal weld solver reference",        # 오픈셀 내부 Weld 솔버 레퍼런스
        "cell_solref": "OpenCell general contact solver reference",           # 오픈셀 일반 접촉 솔버 레퍼런스
        "enable_air_drag": "Enable air drag and viscous forces",             # 공기 저항 및 점성력 활성화
        "enable_air_squeeze": "Enable air squeeze film effect on impact",     # 충돌 시 공기 압축(Squeeze Film) 효과 활성화
        "air_density": "Air density (kg/m^3)",                                # 공기 밀도 (kg/m^3)
        "air_viscosity": "Air dynamic viscosity (Pa.s)",                      # 공기 점성 계수 (Pa.s)
        "air_cd_drag": "Blunt body drag coefficient",                         # 형상 저항 계수 (Drag Coefficient)
        "air_cd_viscous": "Viscous skin friction coefficient",                # 점성 마찰 계수 (Viscous Coeff)
        "air_coef_squeeze": "Squeeze film intensity multiplier",              # 스퀴즈 효과 강도 배율
        "air_squeeze_hmax": "Max height for squeeze film activation (m)",      # 스퀴즈 효과 활성화 최대 높이 (m)
        "air_squeeze_hmin": "Min height for squeeze film activation (m)"       # 스퀴즈 효과 활성화 최소 높이 (m)
    }

    logged_keys = set()
    for cat, keys in categories.items():
        cat_lines = []
        for k in keys:
            if k in config:
                val = config[k]
                desc = descriptions.get(k, "")
                comment = f" # {desc}" if desc else ""
                cat_lines.append(f"  - {k:<25}: {str(val):<20}{comment}")
                logged_keys.add(k)
        
        if cat_lines:
            lines.append(f"\n[{cat}]")
            lines.extend(cat_lines)
    
    # 8. 추가 질량(Aux Masses) 별도 처리
    if "chassis_aux_masses" in config and config["chassis_aux_masses"]:
        lines.append("\n[8. 추가 질량 (Chassis Aux Masses)]")
        for i, aux in enumerate(config["chassis_aux_masses"]):
            name = aux.get('name', f'Aux_{i}')
            lines.append(f"  * {name:<23}: pos={str(aux.get('pos')):<35} # mass={aux.get('mass')}kg, size={aux.get('size')}")
        logged_keys.add("chassis_aux_masses")

    # 9. 시각화 및 기타 (Visual & Misc)
    visual_keys = ["light_main_diffuse", "light_main_ambient", "light_head_diffuse", "light_head_ambient", "light_sub_diffuse"]
    vis_lines = []
    for k in visual_keys:
        if k in config:
            desc = descriptions.get(k, "")
            comment = f" # {desc}" if desc else ""
            vis_lines.append(f"  - {k:<25}: {str(config[k]):<20}{comment}")
            logged_keys.add(k)
    if vis_lines:
        lines.append("\n[9. 시각화 관련 설정 (Visual)]")
        lines.extend(vis_lines)

    # 상기 카테고리에 누락된 기타 설정들 (진짜 기타)
    other_keys = sorted([k for k in config.keys() if k not in logged_keys and not k.startswith("mat_")])
    if other_keys:
        lines.append("\n[10. 기타 보조 설정 (Others)]")
        for k in other_keys:
            desc = descriptions.get(k, "")
            comment = f" # {desc}" if desc else ""
            lines.append(f"  - {k:<25}: {str(config[k]):<20}{comment}")
        
    lines.append("\n" + "="*90 + "\n")
    return "\n".join(lines)

def print_inertia_report(config, title="[Assembly Inertia Report]", logger=print):
    """지정된 설정에 따른 모델의 질량, CoG, MoI 정보를 측정하고 리포트 형식으로 출력합니다."""
    # 실제 XML 파일 생성을 피하기 위해 임시 파일명을 사용하지만, 
    # create_model 내부에서 이미 로깅 로직이 있으므로 이를 활용합니다.
    temp_xml = "temp_inertia_check.xml"
    _, total_mass, cog, moi, details = create_model(temp_xml, config=config, logger=logger)
    return total_mass, cog, moi, details

def save_config_as_py(config, filepath, timestamp):
    """전체 설정을 차후 재현 가능한 Python 딕셔너리 형태의 파일로 저장합니다."""
    import pprint
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f'# MuJoCo Drop Simulation Reproduction Config\n')
        f.write(f'# Generated for Run: rds-{timestamp}\n')
        f.write(f'# Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
        formatted_cfg = pprint.pformat(config, indent=4, sort_dicts=False)
        f.write(f'config = {formatted_cfg}\n')

plt.rcParams.update({'font.size': 8})



def ask_use_viewer():
    """사용자에게 GUI 뷰어(Passive Viewer) 사용 여부를 묻습니다."""
    try:
        prompt = "\n>> MuJoCo GUI 뷰어를 실행하시겠습니까? (Y/n, 기본값=Y): "
        ans = input(prompt).strip().lower()
        
        # [LOGGING] 사용자 응답을 로그에 기록
        log_and_print(f"{prompt.strip()} -> {ans if ans else 'y (default)'}")
        
        if ans == 'n':
            return False
        return True
    except EOFError:
        log_and_print(">> IO Error or Empty input. Defaulting to True.")
        return True


def calc_mujoco_stiffness(solref_str, mass=1.0):
    """
    MuJoCo solref (timeconst, dampratio)를 기반으로 유효 강성 K와 감쇠 C를 계산합니다.
    MuJoCo 공식: 
      K = 1 / (timeconst^2 * dampratio^2)
      C = 2 / timeconst
    물질의 질량(mass)을 곱하여 최종 차원(N/m, Ns/m)을 반환합니다.
    """
    try:
        parts = [float(x) for x in str(solref_str).split()]
        if len(parts) < 2:
            return 0.0, 0.0
        tc, dr = parts[0], parts[1]
        
        # tc 또는 dr 이 0이면 수치적 발산 방지
        if tc <= 0 or dr <= 0:
            return 0.0, 0.0
            
        k_unit = 1.0 / (tc**2 * dr**2)
        c_unit = 2.0 / tc
        
        return k_unit * mass, c_unit * mass
    except Exception:
        return 0.0, 0.0



def apply_aerodynamics(model, data, config):
    """
    공기 저항(Drag, Viscous) 및 지면 근접 시 압축 공기 효과(Squeeze Film)를 계산하여 적용합니다.
    [개선] 특정 한 면만 계산하는 방식에서 모든 하향 면(Downward Faces)을 체크하는 방식으로 변경하여 
    Corner/Edge 낙하 시에도 공기 쿠션 효과가 누락되지 않도록 합니다.
    """
    rho      = config.get('air_density', 1.225)
    mu       = config.get('air_viscosity', 1.81e-5)
    Cd_blunt = config.get('air_cd_drag', 1.05)
    Cd_visc  = config.get('air_cd_viscous', 0.0)
    Csq      = config.get('air_coef_squeeze', 0.0)
    h_sq_max = config.get('air_squeeze_hmax', 0.20)
    h_sq_min = config.get('air_squeeze_hmin', 0.001)
    
    # 적용 대상 바디 (최상위 박스)
    body_name = "BPackagingBox"
    body_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, body_name)
    if body_id == -1:
        return 0.0, 0.0, 0.0

    # 외력 초기화 (MuJoCo는 제어 콜백에서 직접 xfrc_applied를 설정하면 해당 스텝에 반영됨)
    data.xfrc_applied[body_id, :] = 0.0

    # CoM 상태 정보 추출
    pos      = data.xpos[body_id]
    mat      = data.xmat[body_id].reshape(3, 3)
    ang_vel  = data.cvel[body_id, 0:3] 
    lin_vel  = data.cvel[body_id, 3:6] 

    f_sq_total = 0.0
    total_torque = np.zeros(3)

    # [1] Squeeze Film 효과 계산 (모든 면에 대해)
    if config.get('enable_air_squeeze', True) and Csq > 0:
        dims = [config.get('box_w', 2.0), config.get('box_h', 1.4), config.get('box_d', 0.25)]
        # 각 로컬 축(X, Y, Z)에 대하여
        for axis_idx in range(3):
            axis_vec = mat[:, axis_idx]  # 로컬 축의 월드 벡터
            dot_z = axis_vec[2]          # 월드 Z축 투영 성분
            
            if abs(dot_z) < 1e-3: continue # 지면과 수직인 면은 스퀴즈 없음
            
            # 하향 방향의 법선 설정
            local_norm = np.zeros(3)
            local_norm[axis_idx] = -np.sign(dot_z) * (dims[axis_idx] / 2.0)
            
            # 해당 면의 가로/세로 길이 결정
            other_indices = [i for i in range(3) if i != axis_idx]
            u_len = dims[other_indices[0]]
            v_len = dims[other_indices[1]]
            u_vec = mat[:, other_indices[0]]
            v_vec = mat[:, other_indices[1]]
            
            # 형상 계수 계산
            geo_factor = ((u_len * v_len) / (2 * (u_len + v_len))) ** 2
            P_COEF = 0.5 * rho * geo_factor * Csq
            
            face_center_rel = mat @ local_norm
            
            # 격자 적산 (N x N)
            N = 6 # 성능을 위해 약간 조정
            dA = (u_len * v_len) / (N * N)
            grid_steps = np.linspace(-0.5+0.5/N, 0.5-0.5/N, N)
            
            for u in grid_steps:
                for v in grid_steps:
                    rel_p = face_center_rel + (u * u_len) * u_vec + (v * v_len) * v_vec
                    h = pos[2] + rel_p[2]
                    
                    if h_sq_min < h < h_sq_max:
                        # 점 속도 계산
                        p_vel = lin_vel + np.cross(ang_vel, rel_p)
                        vz = p_vel[2]
                        if vz < 0:
                            dF = P_COEF * dA * (vz / h)**2
                            dF = min(dF, 500.0 / (N*N/64)) # 포인트당 캡
                            f_sq_total += dF
                            total_torque += np.cross(rel_p, np.array([0, 0, dF]))

        # 물리 엔진에 힘 적용
        data.xfrc_applied[body_id, 2] += f_sq_total
        data.xfrc_applied[body_id, 3:6] += total_torque

    # [2] 로깅용 에어로다이나믹 추정값 (Drag/Viscous)
    # MuJoCo가 계산하는 값을 근사하여 반환 (시각화용)
    total_area = 2*(dims[0]*dims[1] + dims[1]*dims[2] + dims[2]*dims[0]) if 'dims' in locals() else 5.0
    v_mag = np.linalg.norm(lin_vel)
    f_drag_est = 0.5 * rho * v_mag**2 * Cd_blunt * (total_area/6.0) if config.get('enable_air_drag', True) else 0.0
    f_visc_est = mu * v_mag * Cd_visc * total_area if config.get('enable_air_drag', True) else 0.0

    return f_drag_est, f_visc_est, f_sq_total



def get_body_kinematics(model, data, body_name):
    body_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, body_name)
    if body_id == -1:
        return None
    pos = data.xpos[body_id].copy()
    mat = data.xmat[body_id].reshape(3, 3).copy()
    vel = data.cvel[body_id].copy() # [w_x, w_y, w_z, v_x, v_y, v_z]
    acc = data.cacc[body_id].copy() # [alpha_x, alpha_y, alpha_z, a_x, a_y, a_z]
    return pos, mat, vel, acc

def compute_corners(center_pos, center_mat, box_w, box_h, box_d):
    """지정된 중심점과 회전 행렬을 기반으로 8개 모서리 좌표 역계산"""
    corners_local = []
    for x in [-box_w/2, box_w/2]:
        for y in [-box_h/2, box_h/2]:
            for z in [-box_d/2, box_d/2]:
                corners_local.append(np.array([x, y, z]))
    
    corners_global = []
    for loc in corners_local:
        glob = center_pos + center_mat @ loc
        corners_global.append(glob)
    return np.array(corners_global)

def compute_corner_kinematics(center_pos, center_mat, center_vel, center_acc, box_w, box_h, box_d):
    # center_vel: [wx, wy, wz, vx, vy, vz]
    w = center_vel[0:3]
    v = center_vel[3:6]
    
    # center_acc: [alphax, alphay, alphaz, ax, ay, az]
    alpha = center_acc[0:3]
    a = center_acc[3:6]
    
    corners_local = []
    for x in [-box_w/2, box_w/2]:
        for y in [-box_h/2, box_h/2]:
            for z in [-box_d/2, box_d/2]:
                corners_local.append(np.array([x, y, z]))
                
    results = []
    for loc in corners_local:
        # global offset vector from center
        r = center_mat @ loc
        
        # velocity = v + w x r
        v_corner = v + np.cross(w, r)
        
        # acceleration = a + alpha x r + w x (w x r)
        a_corner = a + np.cross(alpha, r) + np.cross(w, np.cross(w, r))
        
        results.append({
            'pos': center_pos + r,
            'vel': v_corner,
            'acc': a_corner
        })
    return results

def run_simulation(config_or_path, sim_duration=0.5):
    # [NEW] 결과 저장을 위한 타임스탬프 및 출력 경로 초기화 (가장 먼저 수행)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"rds-{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    
    # 전역 로깅 경로 설정
    global _report_file_path
    _report_file_path = os.path.abspath(os.path.join(output_dir, "summary_report.txt"))

    # 시작 시점에 빈 파일을 생성(Overwrite)하여 기록 준비
    with open(_report_file_path, "w", encoding="utf-8") as f:
        pass 

    if isinstance(config_or_path, str):
        # XML 경로가 직접 전달된 경우
        xml_path = config_or_path
        log_and_print(f"\nLoading MuJoCo model from XML: {xml_path}")
        model = mujoco.MjModel.from_xml_path(xml_path)
        config = get_default_config() # fallback
    else:
        # Config 딕셔너리가 전달된 경우 (모델 생성 필요)
        config = get_default_config(config_or_path)
        
        # [NEW] 로깅 디렉토리 생성 직후, 보정 전(Baseline) 관성 상태를 측정하여 요약에 포함할 준비
        baseline_stats = None
        if not config.get("only_generate_xml", False):
            # 보정 전 상태를 확인하기 위해 aux masses가 없는 임시 설정을 만듭니다.
            base_cfg = config.copy()
            if "chassis_aux_masses" in base_cfg:
                base_cfg["chassis_aux_masses"] = []
            
            # 로그 파일이 생성되었으므로, 여기에 Baseline 기록 (터미널 출력은 생략하거나 별도 처리)
            _, b_m, b_c, b_i, _ = create_model("temp_baseline.xml", config=base_cfg, logger=lambda x: None)
            baseline_stats = (b_m, b_c, b_i)

        # Detailed summary of current configuration
        summary_text = format_config_report(config, timestamp)
        
        # [NEW] Baseline 정보를 텍스트 하단에 병합
        if baseline_stats:
            b_m, b_c, b_i = baseline_stats
            extra = []
            extra.append("\n[Baseline Inertia (Pre-Correction)]")
            extra.append(f"  - Baseline Mass            : {b_m:12.4f} kg")
            extra.append(f"  - Baseline CoG             : ({b_c[0]:.4f}, {b_c[1]:.4f}, {b_c[2]:.4f})")
            extra.append(f"  - Baseline MoI (x,y,z)     : ({b_i[0]:.6f}, {b_i[1]:.6f}, {b_i[2]:.6f})")
            summary_text = summary_text.replace("\n" + "="*90 + "\n", "\n".join(extra) + "\n\n" + "="*90 + "\n")

        log_and_print(summary_text)

        config_py_path = os.path.join(output_dir, f"rds-{timestamp}_config.py")
        save_config_as_py(config, config_py_path, timestamp)
        log_and_print(f"  >> Reproducible config script saved to: {os.path.basename(config_py_path)}")

        log_and_print("\nGenerating discrete box model from config...")
        xml_file = "temp_drop_sim.xml"
        xml_str, *_ = create_model(xml_file, config=config, logger=log_and_print)
        xml_abs_path = os.path.abspath(xml_file)
        log_and_print(f"  >> Generated XML: {xml_abs_path}")
        
        model = mujoco.MjModel.from_xml_string(xml_str)
        sim_duration = config.get('sim_duration', sim_duration)
        xml_path = xml_abs_path

    # [NEW] 모델 생성만 수행하고 시뮬레이션은 건너뛰는 옵션 처리
    if config.get("only_generate_xml", False):
        if xml_path and os.path.exists(xml_path):
            shutil.copy(xml_path, os.path.join(output_dir, os.path.basename(xml_path)))
            log_and_print(f"  >> [XML Mode] XML saved to {output_dir}. Skipping simulation.")
        return None
    
    data = mujoco.MjData(model)
    
    # [NEW] 원본 지오메트리 정보 저장 (Reset 시 복구용)
    original_geom_pos = model.geom_pos.copy()
    original_geom_size = model.geom_size.copy()
    original_geom_rgba = model.geom_rgba.copy()
    
    # [INFO] MuJoCo 3.0+ 멀티코어 연산은 XML의 <option npoolthread="N"> 설정을 통해 활성화됩니다.
    nthread = config.get("sim_nthread", 4)
    if nthread > 1:
        log_and_print(f"  >> Multicore requested: {nthread} threads (via npoolthread).")
    
    # Identify bodies for structural metrics
    # Group by component name -> dict of (i,j,k) -> body_id
    components = {}
    nominal_local_pos = {}
    
    for i in range(model.nbody):
        name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_BODY, i)
        if name and name.startswith("b_"):
            parts = name.split("_")
            # 최소 'b_{comp}_{i}_{j}_{k}' 형식을 갖추기 위해 5개 이상의 토큰이 필요함
            if len(parts) >= 5:
                # 인덱스 (i, j, k)는 마지막 3개 세그먼트에서 추출
                try:
                    idx_i = int(parts[-3])
                    idx_j = int(parts[-2])
                    idx_k = int(parts[-1])
                    
                    # 컴포넌트 이름은 첫 'b_'와 인덱스 사이의 모든 세그먼트를 결합하여 소문자로 저장
                    comp_name = "_".join(parts[1:-3]).lower()
                    
                    if comp_name not in components:
                        components[comp_name] = {}
                    components[comp_name][(idx_i, idx_j, idx_k)] = i
                    nominal_local_pos[i] = model.body_pos[i].copy()
                except ValueError:
                    # 인덱스 변환 실패 시 (이산 블록 형식이 아님) 건너뜀
                    continue

    # [REFINED] 전 부품에 대한 실제 격자 인덱스 범위를 GEOM 기반으로 전수 조사
    # Body 이름 전수 조사 시 Single-Body(Weld=False) 부품의 탐색 누락을 방지하기 위함
    comp_max_idxs = {}
    for i in range(model.ngeom):
        g_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, i)
        if g_name and g_name.startswith("g_"):
            parts = g_name.split("_")
            # g_{comp}_{i}_{j}_{k} 형식 (최소 5개 세그먼트)
            if len(parts) >= 5:
                # 'g'와 인덱스들 사이가 컴포넌트 이름
                comp_name = "_".join(parts[1:-3]).lower()
                try:
                    i_idx = int(parts[-3])
                    j_idx = int(parts[-2])
                    k_idx = int(parts[-1])
                    
                    if comp_name not in comp_max_idxs:
                        comp_max_idxs[comp_name] = [0, 0, 0]
                    
                    if i_idx > comp_max_idxs[comp_name][0]: comp_max_idxs[comp_name][0] = i_idx
                    if j_idx > comp_max_idxs[comp_name][1]: comp_max_idxs[comp_name][1] = j_idx
                    if k_idx > comp_max_idxs[comp_name][2]: comp_max_idxs[comp_name][2] = k_idx
                except:
                    pass
    
    # 튜플로 변환하여 확정
    for k in list(comp_max_idxs.keys()):
        comp_max_idxs[k] = tuple(comp_max_idxs[k])
        
    # [VISUALIZATION] 코너(수직 엣지) 블록 색상 변경 (Yellow)
    corner_count = 0
    for comp_name, (nx_max, ny_max, nz_max) in comp_max_idxs.items():
        if "cushion" in comp_name:
            for gid in range(model.ngeom):
                t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, gid)
                if t_name and t_name.lower().startswith(f"g_{comp_name}_"):
                    parts = t_name.split('_')
                    if len(parts) >= 5:
                        try:
                            c_i, c_j = int(parts[-3]), int(parts[-2])
                            if (c_i == 0 or c_i == nx_max) and (c_j == 0 or c_j == ny_max):
                                model.geom_rgba[gid] = [1.0, 1.0, 0.0, 1.0] # Yellow
                                corner_count += 1
                        except: pass
    
    if comp_max_idxs:
        log_and_print(f"  >> [Plasticity] Component Index Ranges: {comp_max_idxs}")
        log_and_print(f"  >> [Plasticity] {corner_count} corner geoms highlighted in Yellow.")

    # [REFINED] Neighbor Discovery for Strain Calculation
    corner_neighbor_pairs = []
    for comp_name, (nx_max, ny_max, nz_max) in comp_max_idxs.items():
        if "cushion" in comp_name:
            comp_geoms = {} # (i,j,k) -> gid
            for gid in range(model.ngeom):
                name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, gid)
                if not name: continue
                if name.lower().startswith("g_") and comp_name in name.lower():
                    parts = name.split('_')
                    if len(parts) >= 5:
                        try:
                            # Use parts from the original name to be safe
                            idx = (int(parts[-3]), int(parts[-2]), int(parts[-1]))
                            comp_geoms[idx] = gid
                        except: pass
            
            # log_and_print(f"  DEBUG: comp={comp_name}, geoms_found={len(comp_geoms)}")
            
            for idx, gid in comp_geoms.items():
                ci, cj, ck = idx
                # Corner block detection (Z-edge corner or face corner)
                if (ci == 0 or ci == nx_max) and (cj == 0 or cj == ny_max):
                    lookups = []
                    # Inward neighbors
                    if ci == 0 and nx_max > 0:       lookups.append(((1, cj, ck), 0))
                    if ci == nx_max and nx_max > 0:  lookups.append(((nx_max-1, cj, ck), 0))
                    if cj == 0 and ny_max > 0:       lookups.append(((ci, 1, ck), 1))
                    if cj == ny_max and ny_max > 0:  lookups.append(((ci, ny_max-1, ck), 1))
                    if ck == 0 and nz_max > 0:       lookups.append(((ci, cj, 1), 2))
                    if ck == nz_max and nz_max > 0:  lookups.append(((ci, cj, nz_max-1), 2))
                    
                    for n_idx, axis in lookups:
                        if n_idx in comp_geoms:
                            nid = comp_geoms[n_idx]
                            b1 = model.geom_bodyid[gid]
                            b2 = model.geom_bodyid[nid]
                            # Initial distance in box local frame
                            d_init = np.linalg.norm(model.body_pos[b1] - model.body_pos[b2])
                            if d_init > 0.001:
                                corner_neighbor_pairs.append((gid, nid, axis, d_init))

    if corner_neighbor_pairs:
        log_and_print(f"  >> [Plasticity] Initialized {len(corner_neighbor_pairs)} neighbor pairs for strain calculation.")
    else:
        log_and_print(f"  >> [Plasticity WARNING] No neighbor pairs found! comp_max_idxs={comp_max_idxs}")


    # Time setup
    dt = model.opt.timestep
    duration = config.get("sim_duration", 1.0)
    steps = int(duration / dt)
    
    time_history = []           # 시뮬레이션 경과 시간 (Simulation Time)
    z_hist = []                  # 전체 조립체의 수직(Z) 방향 높이 (Height)
    pos_hist = []                # 전체 조립체의 6자유도 위치 성분 (Root Position)
    vel_hist = []                # 전체 조립체의 6자유도 속도 성분 (Root Velocity)
    acc_hist = []                # 전체 조립체의 6자유도 가속도 성분 (Root Acceleration)
    cog_pos_hist = []            # 전체 조립체의 무게중심 전역 위치 (Subtree CoM Position)
    cog_vel_hist = []            # 전체 조립체의 무게중심 전역 속도 (Subtree CoM Velocity)
    cog_acc_hist = []            # 전체 조립체의 무게중심 전역 가속도 (Subtree CoM Acceleration)
    corner_pos_hist = []         # 박스 외곽 8개 모서리의 전역 위치 (8 Corners World Position)
    corner_vel_hist = []         # 박스 외곽 8개 모서리의 전역 속도 (8 Corners World Velocity)
    corner_acc_hist = []         # 박스 외곽 8개 모서리의 전역 가속도 (8 Corners World Acceleration)
    ground_impact_hist = []      # 지면과의 접촉 및 충격력 크기 (Ground Impact Force Magnitude)
    air_drag_hist = []           # 계산된 공기 형상 저항 추정치 (Estimated Air Drag)
    air_viscous_hist = []        # 계산된 공기 점성 마찰 저항 추정치 (Estimated Air Viscous)
    air_squeeze_hist = []        # 지면 낙하 직전 공기 압축(Squeeze Film) 저항 (Air Squeeze Force)
    
    # [NEW] 영구 변형(Plastic Deformation) 및 물리 계수 사전 계산
    enable_plasticity = config.get("enable_plasticity", False)
    plasticity_ratio = config.get("plasticity_ratio", 0.3)
    yield_stress_pa = config.get("cush_yield_stress", 0.01) * 1e6 # MPa -> Pa
    yield_strain = config.get("cush_yield_strain", 0.1) # 10% compression threshold

    # Metric histories per component by row (j) and individual blocks
    # [NEW] MuJoCo 물리 파라미터 (Stiffness K, Damping C) 사전 계산 및 출력
    cush_tc = config.get("cush_weld_solref_timec", config.get("cush_solref_timec", 0.02))
    cush_dr = config.get("cush_weld_solref_damprr", config.get("cush_solref_damprr", 1.0))
    cush_weld_solref = f"{cush_tc} {cush_dr}"
    
    # 쿠션 블록 한 개의 질량 (에너지 계산용)
    cush_div = config.get("cush_div", [5, 4, 3])
    num_cush_blocks = np.prod(cush_div)
    m_cush_block = config.get("mass_cushion", 1.0) / num_cush_blocks
    
    k_cush, c_cush = calc_mujoco_stiffness(cush_weld_solref, m_cush_block)
    
    # [NEW] 탄성 계수 (Young's Modulus, E) 근사 계산
    # E = (K * L) / A  (K: 강성, L: 블록 두께, A: 단면적)
    cush_w = config.get("box_w", 2.0) - 2*config.get("box_thick", 0.01)
    cush_h = config.get("box_h", 1.4) - 2*config.get("box_thick", 0.01)
    cush_d = config.get("box_d", 0.25) - 2*config.get("box_thick", 0.01)

    avg_dx = cush_w / cush_div[0]
    avg_dy = cush_h / cush_div[1]
    avg_dz = cush_d / cush_div[2]
    
    # 단면적 (xy 평면 기준) 및 두께 (z 방향)
    area_avg = avg_dx * avg_dy
    E_estimated = (k_cush * avg_dz) / area_avg if area_avg > 0 else 0.0
    E_mpa = E_estimated / 1e6 # Pa -> MPa
    E_kpa = E_estimated / 1e3 # Pa -> kPa

    # 에너지 계산용 프록시 강성을 실제 물리값으로 업데이트
    k_spring_proxy = k_cush

    # 바닥 및 테이프 정보 (참조용 전역 질량 기준)
    total_mass = float(np.sum(model.body_mass))
    k_ground, c_ground = calc_mujoco_stiffness(config.get("ground_solref", "0.01 1.0"), total_mass)
    k_tape,   c_tape   = calc_mujoco_stiffness(config.get("tape_solref", "0.005 1.0"), config.get("mass_occ", 0.1))

    log_and_print("\n" + "="*80)
    log_and_print(f"[MuJoCo Solver Parameters (Calculated K & C)]")
    log_and_print(f" - Ground (Global) : K = {k_ground:10.1f} N/m, C = {c_ground:8.1f} Ns/m")
    log_and_print(f"   (Used: mass={total_mass:.3f}kg, solref={config.get('ground_solref', '0.01 1.0')})")
    
    log_and_print(f" - Tape (Global)   : K = {k_tape:10.1f} N/m, C = {c_tape:8.1f} Ns/m")
    log_and_print(f"   (Used: mass={config.get('mass_occ', 0.1):.3f}kg, solref={config.get('tape_solref', '0.005 1.0')})")
    
    log_and_print(f" - Cushion (Weld)  : K = {k_cush:10.1f} N/m, C = {c_cush:8.1f} Ns/m (per block)")
    log_and_print(f"   (Used: block_mass={m_cush_block:.6f}kg, solref={cush_weld_solref})")
    log_and_print(f"   (Geom: Avg_Area={area_avg:.6f} m^2, Avg_Depth={avg_dz:.4f} m)")
    log_and_print(f"   >> Est. Young's Modulus (E): {E_mpa:.4e} MPa ({E_kpa:.2f} kPa)")
    if enable_plasticity:
        yield_pressure = config.get('cush_yield_pressure', 0.0) * 1000.0 # kPa to Pa
        log_and_print(f"   >> Yield Thresholds: Stress = {config.get('cush_yield_stress', 0.01)} MPa, Strain = {yield_strain:.2f}, Pressure = {config.get('cush_yield_pressure', 0.0)} kPa")
    log_and_print("="*80 + "\n")

    metrics = {}
    for comp in components:
        metrics[comp] = {}
        metrics[comp]['all_blocks_angle'] = {}
        metrics[comp]['block_nominals'] = {}
        metrics[comp]['block_nominal_mats'] = {} # [NEW] 초기 오리엔테이션 기준값 저장용
        metrics[comp]['total_distortion'] = [] # Overall deformation index (RMS of angles)
        for block_idx in components[comp]:
            metrics[comp]['all_blocks_angle'][block_idx] = []
            metrics[comp]['block_nominals'][block_idx] = nominal_local_pos[components[comp][block_idx]]
            metrics[comp]['block_nominal_mats'][block_idx] = None
            
        # [NEW] Corner Plasticity tracking (12 corner blocks)
        metrics[comp]['corner_hists'] = {}
        # Find corner blocks for this component (bcushion에 대해서만 데이터 트래킹 및 그래프 생성 수행)
        if comp == 'bcushion':
            nx_max, ny_max, nz_max = comp_max_idxs.get(comp, (0,0,0))
            for block_idx in components[comp]:
                ci, cj, ck = block_idx
                if (ci == 0 or ci == nx_max) and (cj == 0 or cj == ny_max):
                    gid = -1
                    # Finding gid from component name and indices
                    g_prefix = f"g_{comp}_"
                    for g_idx in range(model.ngeom):
                        g_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, g_idx)
                        if g_name and g_name.lower().startswith(g_prefix) and g_name.endswith(f"_{ci}_{cj}_{ck}"):
                            gid = g_idx
                            break
                    if gid != -1:
                        metrics[comp]['corner_hists'][gid] = {
                            'strain': [], 'pressure': [], 'disp': [], 'plastic': [], 'name': g_name
                        }

        # Find unique j indices
        j_idx = set([k[1] for k in components[comp].keys()])
        for j in j_idx:
            metrics[comp][j] = {
                'bending': [], 'twist': [], 'energy': [],
                'loc_b': [], 'loc_t': [], 'loc_e': []
            }
            
    log_and_print(f"Starting simulation for {duration} seconds with {steps} steps...")
    
    # k_spring_proxy is now calculated from MuJoCo parameters above
    
    prev_vel_z = 0.0
    
    mujoco.mj_forward(model, data)
    
    # [NEW] 영구 변형(Plastic Deformation) 로직용 상태 저장소
    geom_state_tracker = {}
    
    def apply_plastic_deformation_v1():
        if not enable_plasticity: return
        current_penetrations = {}
        geom_hits = {} # gid -> {'max_p', 'sum_f', 'local_n', 'parts'}
        for i in range(data.ncon):
            con = data.contact[i]
            g1_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, con.geom1)
            g2_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, con.geom2)
            target_geom = -1
            if g1_name and "ground" in g1_name.lower(): target_geom = con.geom2
            elif g2_name and "ground" in g2_name.lower(): target_geom = con.geom1
            if target_geom != -1:
                t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, target_geom)
                if t_name and "cushion" in t_name.lower(): 
                    parts = t_name.split('_')
                    if len(parts) >= 5:
                        gid = target_geom
                        if gid not in geom_hits:
                            geom_hits[gid] = {'max_p': 0.0, 'sum_f': 0.0, 'local_n': np.zeros(3), 'parts': parts, 'name': t_name}
                        force_vec = np.zeros(6)
                        mujoco.mj_contactForce(model, data, i, force_vec)
                        geom_hits[gid]['sum_f'] += abs(force_vec[0])
                        pen = -con.dist
                        if pen > geom_hits[gid]['max_p']:
                            geom_hits[gid]['max_p'] = pen
                            body_id = model.geom_bodyid[gid]
                            nw = con.frame[:3]
                            geom_hits[gid]['local_n'] = data.xmat[body_id].reshape(3,3).T @ nw
        for gid, hit in geom_hits.items():
            parts = hit['parts']
            try:
                c_i = int(parts[-3]); c_j = int(parts[-2])
                comp_name = "_".join(parts[1:-3]).lower()
                nx_max, ny_max, nz_max = comp_max_idxs.get(comp_name, (0,0,0))
                if (c_i == 0 or c_i == nx_max) and (c_j == 0 or c_j == ny_max):
                    sz = model.geom_size[gid]
                    areas = [sz[1]*sz[2], sz[0]*sz[2], sz[0]*sz[1]]
                    g_area = 4.0 * min(areas) 
                    local_n = hit['local_n']
                    ma = int(np.argmax(np.abs(local_n))) 
                    current_penetrations[gid] = hit['max_p']
                    pressure = hit['sum_f'] / g_area if g_area > 0 else 0
                    if pressure > yield_stress_pa:
                        if hit['max_p'] > 1e-6:
                            if gid not in geom_state_tracker:
                                geom_state_tracker[gid] = {'max_p': 0.0, 'major_axis': ma}
                                if pressure > 0:
                                    log_and_print(f"  [Plasticity] Corner Activated(v1): {hit['name']} (Pressure: {pressure/1e3:.1f}kPa, Axis: {ma})")
                            if hit['max_p'] > geom_state_tracker[gid]['max_p']:
                                geom_state_tracker[gid]['max_p'] = hit['max_p']
                                geom_state_tracker[gid]['major_axis'] = ma
                                geom_state_tracker[gid]['last_pressure'] = pressure
            except: pass
        for gid, state in geom_state_tracker.items():
            curr_p = current_penetrations.get(gid, 0.0)
            if state['max_p'] > 0.0001 and curr_p < state['max_p']:
                delta_p = state['max_p'] - curr_p
                deformation = delta_p * plasticity_ratio
                if deformation > 1e-6:
                    body_id = model.geom_bodyid[gid]
                    b_pos = model.body_pos[body_id]
                    inward_dir = -np.sign(b_pos)
                    for i_ax in range(3):
                        if abs(inward_dir[i_ax]) < 0.1: inward_dir[i_ax] = -1.0
                    major_axis = state.get('major_axis', 2)
                    model.geom_size[gid][major_axis] = max(0.001, model.geom_size[gid][major_axis] - (deformation/2.0))
                    shift_vec = np.zeros(3)
                    shift_vec[major_axis] = inward_dir[major_axis] * (deformation/2.0)
                    model.geom_pos[gid] += shift_vec
                    total_shrink = original_geom_size[gid][major_axis] - model.geom_size[gid][major_axis]
                    color_scale = min(1.0, total_shrink / 0.005)
                    model.geom_rgba[gid][0:3] = [0.5*(1-color_scale), 0.2*(1-color_scale), 0.6+0.4*color_scale]
                    state['max_p'] = curr_p
                    t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, gid)
                    p_val = state.get('last_pressure', 0.0)
                    t_size = original_geom_size[gid][major_axis]
                    curr_strain = curr_p / t_size if t_size > 0 else 0
                    if p_val >= yield_pressure and curr_strain >= yield_strain:
                        log_and_print(f"  [Plasticity] {t_name} Deforming(v1): -{total_shrink*1000:.1f}mm (Axis: {major_axis}, Pressure: {p_val/1e3:.1f}kPa)")
            if curr_p > state['max_p']: state['max_p'] = curr_p

    def apply_plasticity():
        # [NEW] Pre-calculate contact pressures for all cushion geoms
        current_geom_pressures = {}
        if enable_plasticity:
            for i in range(data.ncon):
                con = data.contact[i]
                for gid in [con.geom1, con.geom2]:
                    name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, gid)
                    if name and "cushion" in name.lower():
                        if gid not in current_geom_pressures:
                            current_geom_pressures[gid] = 0.0
                        force_vec = np.zeros(6)
                        mujoco.mj_contactForce(model, data, i, force_vec)
                        sz = model.geom_size[gid]
                        g_area = 4.0 * min(sz[0]*sz[1], sz[0]*sz[2], sz[1]*sz[2])
                        if g_area > 0:
                            current_geom_pressures[gid] += abs(force_vec[0]) / g_area

        algo = config.get("plasticity_algorithm", 2)
        if algo == 1:
            apply_plastic_deformation_v1() # v1 still uses hits internally
        else:
            apply_plastic_deformation_v2(current_geom_pressures)
            
        # [NEW] Record corner histories (12 points) for plotting and export
        for comp in components:
            if 'corner_hists' in metrics[comp]:
                for cid, hist in metrics[comp]['corner_hists'].items():
                    # Get instantaneous strain & disp from neighbor distance
                    max_s, max_d = 0.0, 0.0
                    for _cid, _nid, _ax, _d0 in corner_neighbor_pairs:
                        if _cid == cid:
                            b1, b2 = model.geom_bodyid[_cid], model.geom_bodyid[_nid]
                            d_curr = np.linalg.norm(data.xpos[b1] - data.xpos[b2])
                            s = (_d0 - d_curr) / _d0 if _d0 > 0 else 0
                            if s > max_s: max_s = s
                            if (_d0 - d_curr) > max_d: max_d = (_d0 - d_curr)
                    
                    hist['strain'].append(max_s)
                    hist['pressure'].append(current_geom_pressures.get(cid, 0.0))
                    hist['disp'].append(max(0, max_d))
                    
                    # Permanent deformation (Total shrinkage)
                    if cid in geom_state_tracker:
                        ma = geom_state_tracker[cid].get('major_axis', 2)
                        shrink = original_geom_size[cid][ma] - model.geom_size[cid][ma]
                    else:
                        shrink = np.max(original_geom_size[cid] - model.geom_size[cid])
                    hist['plastic'].append(max(0, shrink))

    def apply_plastic_deformation_v2(current_geom_pressures):
        if not enable_plasticity: return
        # current_geom_pressures is passed from apply_plasticity

        corner_activation = {} 
        for cid, nid, axis, d0 in corner_neighbor_pairs:
            b1, b2 = model.geom_bodyid[cid], model.geom_bodyid[nid]
            d_curr = np.linalg.norm(data.xpos[b1] - data.xpos[b2])
            
            strain = (d0 - d_curr) / d0 if d0 > 0 else 0
            pressure = current_geom_pressures.get(cid, 0.0)
            
            # [REFINED] Dual-Trigger Logic: Must exceed both strain and pressure thresholds
            if strain > yield_strain and pressure >= yield_pressure:
                p_val = max(0, d0 - d_curr)
                if cid not in corner_activation or strain > corner_activation[cid]['strain']:
                    corner_activation[cid] = {'strain': strain, 'axis': axis, 'p_val': p_val}
        
        # 2. 상태 추적기 업데이트 (Activation)
        for cid, act in corner_activation.items():
            if cid not in geom_state_tracker:
                geom_state_tracker[cid] = {'max_p': 0.0, 'major_axis': act['axis']}
                p_val_act = current_geom_pressures.get(cid, 0.0)
                if p_val_act > 0:
                    t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, cid)
                    log_and_print(f"  [Plasticity] Strain Activated: {t_name} (Strain: {act['strain']:.2f}, Axis: {act['axis']}, Pressure: {p_val_act/1e3:.1f}kPa)")
            
            if act['p_val'] > geom_state_tracker[cid]['max_p']:
                geom_state_tracker[cid]['max_p'] = act['p_val']
                geom_state_tracker[cid]['major_axis'] = act['axis']
                if cid in current_geom_pressures:
                    geom_state_tracker[cid]['last_pressure'] = current_geom_pressures[cid]

        # 3. 실시간 소성 변형 적용
        for cid, state in geom_state_tracker.items():
            ma = state['major_axis']
            d0_ma, d_curr_ma = 0.05, 0.05
            for _cid, _nid, _axis, _d0 in corner_neighbor_pairs:
                if _cid == cid and _axis == ma:
                    d0_ma = _d0
                    d_curr_ma = np.linalg.norm(data.xpos[model.geom_bodyid[_cid]] - data.xpos[model.geom_bodyid[_nid]])
                    break
            
            curr_p = max(0, d0_ma - d_curr_ma)
            
            if state['max_p'] > 0.0005 and curr_p < state['max_p']:
                delta_p = state['max_p'] - curr_p
                deformation = delta_p * plasticity_ratio
                if deformation > 1e-6:
                    body_id = model.geom_bodyid[cid]
                    b_pos = model.body_pos[body_id]
                    inward_dir = -np.sign(b_pos)
                    for i_ax in range(3):
                        if abs(inward_dir[i_ax]) < 0.1: inward_dir[i_ax] = -1.0
                    
                    major_axis = state['major_axis']
                    half_shrink = deformation / 2.0
                    shift_amount = deformation / 2.0
                    
                    model.geom_size[cid][major_axis] = max(0.001, model.geom_size[cid][major_axis] - half_shrink)
                    shift_vec = np.zeros(3)
                    shift_vec[major_axis] = inward_dir[major_axis] * shift_amount
                    model.geom_pos[cid] += shift_vec
                    
                    total_shrink = original_geom_size[cid][major_axis] - model.geom_size[cid][major_axis]
                    color_scale = min(1.0, total_shrink / 0.005)
                    model.geom_rgba[cid][0] = 0.5 * (1.0 - color_scale) # Red component changes
                    state['max_p'] = curr_p
                    t_name = mujoco.mj_id2name(model, mujoco.mjtObj.mjOBJ_GEOM, cid)
                    p_val = state.get('last_pressure', 0.0)
                    curr_strain = curr_p / d0_ma if d0_ma > 0 else 0
                    if p_val >= yield_pressure and curr_strain >= yield_strain:
                        log_and_print(f"  [Plasticity] {t_name} Deforming(v2): -{total_shrink*1000:.1f}mm (Strain: {curr_strain:.2f}, Pressure: {p_val/1e3:.1f}kPa)")
            
            if curr_p > state['max_p']: 
                state['max_p'] = curr_p
                if cid in current_geom_pressures:
                    state['last_pressure'] = current_geom_pressures[cid]
    # [NEW] Control Callback 등록 (mjcb_control)
    def mjcb_aerodynamics(model, data):
        apply_aerodynamics(model, data, config)
    mujoco.set_mjcb_control(mjcb_aerodynamics)
    
    # [NEW] Keyboard Control State & Callback (Reference: boxmotionsim_v0_0_2)
    class ControlState:
        def __init__(self):
            self.paused = False
            self.step_request = False
            self.reset_request = False
            self.quit_request = False
    
    ctrl = ControlState()
    
    def key_callback(keycode):
        if keycode == 32: # Space
            ctrl.paused = not ctrl.paused
            log_and_print(f"   >> [VIEWER] {'PAUSED' if ctrl.paused else 'RUNNING'}")
        elif keycode == 262: # Right Arrow
            ctrl.step_request = True
        elif keycode == 256: # ESC
            ctrl.quit_request = True
            log_and_print("   >> [VIEWER] Quit requested.")
        elif keycode == 259 or keycode == 82: # Backspace (259) or R (82)
            ctrl.reset_request = True
    
    use_viewer = config.get("use_viewer", False)
    viewer = None
    if use_viewer:
        viewer = mujoco.viewer.launch_passive(model, data, key_callback=key_callback)
        log_and_print("\n" + "-"*40)
        log_and_print("🎮 [Passive Viewer Interactive Controls]")
        log_and_print(" - [Space]     : Pause / Resume")
        log_and_print(" - [Right]     : Single Step (when paused)")
        log_and_print(" - [Backspace] : Reset Simulation")
        log_and_print(" - [Esc]       : Stop & Show Results")
        log_and_print("-"*40 + "\n")

    # [NEW] 시뮬레이션 진행 상황 테이블 헤더 출력
    log_and_print("-" * 65)
    log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
    log_and_print("-" * 65)

    start_real_time = time.time()
    last_reported_interval = -1
    last_report_step = 0
    last_report_real_time = start_real_time
    report_count = 0
    
    step = 0
    while step < steps:
        if ctrl.quit_request:
            break
            
        if ctrl.reset_request:
            log_and_print("   >> [VIEWER] Resetting simulation...")
            mujoco.mj_resetData(model, data)
            
            # [NEW] 지오메트리 정보 복구 (Plasticity 초기화)
            model.geom_pos[:] = original_geom_pos
            model.geom_size[:] = original_geom_size
            model.geom_rgba[:] = original_geom_rgba
            geom_state_tracker.clear()
            
            # Re-initialize histories
            time_history.clear(); z_hist.clear(); pos_hist.clear(); vel_hist.clear(); acc_hist.clear()
            cog_pos_hist.clear(); cog_vel_hist.clear(); cog_acc_hist.clear()
            ground_impact_hist.clear(); air_drag_hist.clear(); air_viscous_hist.clear(); air_squeeze_hist.clear()
            corner_pos_hist.clear(); corner_vel_hist.clear(); corner_acc_hist.clear()
            for c in metrics:
                for k in metrics[c]:
                    if isinstance(metrics[c][k], dict):
                        for b in metrics[c][k]: 
                            if isinstance(metrics[c][k][b], list):
                                metrics[c][k][b].clear()
                    elif isinstance(metrics[c][k], list):
                        metrics[c][k].clear()
            step = 0
            ctrl.reset_request = False
            ctrl.paused = True
            
            # Reset table reporting
            last_reported_interval = -1
            start_real_time = time.time()
            last_report_step = 0
            last_report_real_time = start_real_time
            report_count = 0
            log_and_print("-" * 65)
            log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
            log_and_print("-" * 65)
            continue

        if not ctrl.paused or ctrl.step_request:
            try:
                mujoco.mj_step(model, data)
                apply_plasticity()  # 설정에 따른 소성 변형 알고리즘(v1/v2) 차등 적용
                ctrl.step_request = False
                step += 1
            except Exception as e:
                import sys
                sys.stderr.write(f"\n[Drop Sim Error] Simulation unstable: {e}\n")
                break
        else:
            # Paused state: Just sync viewer and sleep a bit to save CPU
            if viewer:
                viewer.sync()
            time.sleep(0.01)
            continue

        if viewer and viewer.is_running():
            viewer.sync()
            
        # [INFO] Progress Reporting (Table Row)
        sim_interval = int(data.time / 0.05)
        if sim_interval > last_reported_interval:
            # Interval FPS calculation
            current_real_time = time.time()
            real_elapsed = current_real_time - start_real_time
            interval_elapsed = current_real_time - last_report_real_time
            if interval_elapsed > 0:
                fps = (step - last_report_step) / interval_elapsed
            else:
                fps = 0.0
            
            log_and_print(f"| {step:8d} | {data.time:12.3f} | {real_elapsed:13.2f} | {fps:8.1f} |")
            
            last_reported_interval = sim_interval
            last_report_step = step
            last_report_real_time = current_real_time
            report_count += 1
            
            # [NEW] 30개 출력마다 열 제목 재출력
            if report_count > 0 and report_count % 30 == 0:
                log_and_print("-" * 65)
                log_and_print(f"| {'Step':^8} | {'Sim Time (s)':^12} | {'Real Time (s)':^13} | {'FPS':^8} |")
                log_and_print("-" * 65)
            
        time_history.append(data.time)
        
        # 1. Root Kinematics (BPackagingBox itself is just a massless container)
        pos, mat, vel, acc = get_body_kinematics(model, data, "BPackagingBox")
        z_hist.append(pos[2])
        pos_hist.append(pos)
        vel_hist.append(vel)
        acc_hist.append(acc)
        
        root_id = mujoco.mj_name2id(model, mujoco.mjtObj.mjOBJ_BODY, "BPackagingBox")
        if root_id != -1:
            cog_pos_hist.append(data.subtree_com[root_id].copy())
            cog_vel_hist.append(data.subtree_linvel[root_id].copy())
        else:
            cog_pos_hist.append(pos)
            cog_vel_hist.append(vel[3:6])
        cog_acc_hist.append(acc[3:6]) # Subtree acceleration in MuJoCo requires numerical differentiation, so proxy by root acceleration
        
        corners = compute_corner_kinematics(pos, mat, vel, acc, config["box_w"], config["box_h"], config["box_d"])
        corner_pos_hist.append([c['pos'] for c in corners])
        corner_vel_hist.append([c['vel'] for c in corners])
        corner_acc_hist.append([c['acc'] for c in corners])
        
        # Ground Impact (Sum of normal forces from worldbody contacts)
        ground_f = 0.0
        for i_con in range(data.ncon):
            contact = data.contact[i_con]
            body1 = model.geom_bodyid[contact.geom1]
            body2 = model.geom_bodyid[contact.geom2]
            if body1 == 0 or body2 == 0:
                forces = np.zeros(6, dtype=np.float64)
                mujoco.mj_contactForce(model, data, i_con, forces)
                ground_f += abs(forces[0])
        ground_impact_hist.append(ground_f)
        
        # Air Resistance Logging (Already applied via callback)
        f_drag, f_viscous, f_squeeze = apply_aerodynamics(model, data, config)
        
        air_drag_hist.append(f_drag)
        air_viscous_hist.append(f_viscous)
        air_squeeze_hist.append(f_squeeze)
        
        # G-Force is now calculated purely from root kinematic differences after simulation
            
        # 2. Component Structural Metrics
        mat_inv = mat.T
        for comp, blocks in components.items():
            # Gather current local positions
            current_local = {}
            for (curr_i, curr_j, curr_k), b_id in blocks.items():
                g_pos = data.xpos[b_id]
                g_mat = data.xmat[b_id].reshape(3, 3)
                rel_pos = g_pos - pos
                l_pos = mat_inv @ rel_pos
                current_local[(curr_i, curr_j, curr_k)] = l_pos
                
                # [REFINED] Calculate angular deformation (Deviation from Initial Orientation)
                # 단순 Root와의 각도 차이가 아니라, 초기(t=0)의 상대적 각도 대비 얼마나 '비틀렸는지'를 계산합니다.
                rot_rel = mat_inv @ g_mat
                if metrics[comp]['block_nominal_mats'][(curr_i, curr_j, curr_k)] is None:
                    metrics[comp]['block_nominal_mats'][(curr_i, curr_j, curr_k)] = rot_rel.copy()
                
                initial_rot_rel = metrics[comp]['block_nominal_mats'][(curr_i, curr_j, curr_k)]
                # Distortion measurement relative to initial configuration
                # distortion = R_rel_init^T * R_rel_curr
                rot_dist = initial_rot_rel.T @ rot_rel
                tr = np.clip(np.trace(rot_dist), -1.0, 3.0)
                theta = np.arccos((tr - 1.0) / 2.0)
                block_angle_deg = np.degrees(theta)
                metrics[comp]['all_blocks_angle'][(curr_i, curr_j, curr_k)].append(block_angle_deg)
            
            # [NEW] Total Distortion Index (TDI) Calculation
            # SDI/TDI = RMS of all block rotation angles
            all_angles = [metrics[comp]['all_blocks_angle'][b_idx][-1] for b_idx in blocks]
            if all_angles:
                tdi_val = np.sqrt(np.mean(np.array(all_angles)**2))
                metrics[comp]['total_distortion'].append(tdi_val)
            else:
                metrics[comp]['total_distortion'].append(0.0)
                
            # Process by row (j)
            j_indices = sorted(list(set([k[1] for k in blocks.keys()])))
            j_mid = j_indices[len(j_indices)//2] if len(j_indices) > 0 else 0
            
            # Mid row slope for twist reference
            mid_row_x = []
            mid_row_z = []
            for (curr_i, curr_j, curr_k), l_pos in current_local.items():
                if curr_j == j_mid:
                    mid_row_x.append(l_pos[0])
                    mid_row_z.append(l_pos[2])
            mid_slope = 0.0
            if len(mid_row_x) > 1:
                mid_slope = np.polyfit(mid_row_x, mid_row_z, 1)[0]
                
            for j in j_indices:
                row_x = []
                row_z = []
                energies = {}
                dz_vals = {}
                
                # [NEW] 현 컴포넌트의 격자 범위 획득
                nx_max, ny_max, nz_max = comp_max_idxs.get(comp, (0,0,0))

                for (curr_i, curr_j, curr_k), l_pos in current_local.items():
                    if curr_j == j:
                        # [NEW] Cushion인 경우 코너 단일 블록은 구조적 굽힘/비틂에서 제외하여 가독성 향상
                        is_corner_block = (curr_i == 0 or curr_i == nx_max) and (curr_j == 0 or curr_j == ny_max)
                        if "cushion" in comp and is_corner_block:
                            continue
                        
                        b_id = blocks[(curr_i, curr_j, curr_k)]
                        nom_pos = nominal_local_pos[b_id]
                        delta = l_pos - nom_pos
                        
                        # Energy proxy: 0.5 * k * dx^2
                        energy = 0.5 * k_spring_proxy * np.sum(delta**2)
                        energies[(curr_i, curr_k)] = energy
                        
                        row_x.append(l_pos[0])
                        row_z.append(l_pos[2])
                        dz_vals[(curr_i, curr_k)] = delta[2]
                
                # Bending: Max Z deflection angle approximation
                bending_deg = 0.0
                loc_b = "N/A"
                if len(dz_vals) > 1:
                    max_idx = max(dz_vals, key=dz_vals.get)
                    min_idx = min(dz_vals, key=dz_vals.get)
                    dz_diff = dz_vals[max_idx] - dz_vals[min_idx]
                    dx_diff = max(row_x) - min(row_x) if max(row_x) != min(row_x) else 1e-6
                    bending_deg = np.degrees(np.arctan(abs(dz_diff / dx_diff)))
                    loc_b = f"{max_idx[0]}_{j}_{max_idx[1]}"
                    
                # Twisting: relative slope diff to mid row
                twist_deg = 0.0
                loc_t = f"*_{j}_*"
                if len(row_x) > 1:
                    slope = np.polyfit(row_x, row_z, 1)[0]
                    twist_deg = np.degrees(np.arctan(abs(slope - mid_slope)))
                    
                # Energy Peak
                peak_energy = 0.0
                loc_e = "N/A"
                if energies:
                    max_e_idx = max(energies, key=energies.get)
                    peak_energy = energies[max_e_idx]
                    loc_e = f"{max_e_idx[0]}_{j}_{max_e_idx[1]}"
                    
                metrics[comp][j]['bending'].append(bending_deg)
                metrics[comp][j]['twist'].append(twist_deg)
                metrics[comp][j]['energy'].append(peak_energy)
                
                # Store locations only if it's a new max (to retrieve easily later, simplified below)
                metrics[comp][j]['loc_b'].append(loc_b)
                metrics[comp][j]['loc_t'].append(loc_t)
                metrics[comp][j]['loc_e'].append(loc_e)

    log_and_print("-" * 65)
    log_and_print("Simulation completed.")
    
    # Calculate global peak G over the differentiated root trajectory
    # Double differentiation of Z position for absolute shock magnitude
    z_array = np.array(z_hist)
    if len(z_array) > 1:
        v_z = np.gradient(z_array, dt)
        a_z = np.gradient(v_z, dt)
        # The actual physics simulation tracks relative to earth freefall, so we pull max impact
        root_acc_history = np.abs(a_z) / 9.81
        max_g_force = float(np.max(root_acc_history))
    else:
        max_g_force = 0.0
    # Generate Terminal Summary Report and save to file
    # [Changed] log_and_print helper is now defined earlier and persistent.

    log_and_print("-" * 50)
    log_and_print(f"Peak Assembly G-Force: {max_g_force:.2f} G")
    log_and_print("-" * 50)

    # [NEW] Generate Final Summary Performance Table in Console & Log
    log_and_print("\n" + "="*95)
    log_and_print(f"  [ MuJoCo DROP TEST - FINAL PERFORMANCE SUMMARY ]")
    log_and_print("="*95)
    log_and_print(f"{'Component':<20} | {'Root Max G':<12} | {'Max Bend(deg)':<15} | {'Max Twist(deg)':<15} | {'Max Plastic(mm)':<15}")
    log_and_print("-" * 95)
    
    log_and_print(f"{'Total Assembly':<20} | {max_g_force:<12.2f} | {'-':<15} | {'-':<15} | {'-':<15}")
    
    for comp in sorted(components.keys()):
        if len(components[comp]) <= 1: continue
        # Calculate peaks for this component
        max_b, max_t, max_p = 0.0, 0.0, 0.0
        j_keys = [k for k in metrics[comp].keys() if isinstance(k, int)]
        for j in j_keys:
            if metrics[comp][j]['bending']: max_b = max(max_b, max(metrics[comp][j]['bending']))
            if metrics[comp][j]['twist']:   max_t = max(max_t, max(metrics[comp][j]['twist']))
        if 'corner_hists' in metrics[comp]:
            for cid in metrics[comp]['corner_hists']:
                if metrics[comp]['corner_hists'][cid]['plastic']:
                    max_p = max(max_p, max(metrics[comp]['corner_hists'][cid]['plastic']))
        
        log_and_print(f"{comp:<20} | {'-':<12} | {max_b:<15.2f} | {max_t:<15.2f} | {max_p*1000:<15.2f}")
    log_and_print("="*95 + "\n")

    for comp, j_data in metrics.items():
        # [NEW] 리지드 바디(Rigid Body) 및 단일 블록(Aux Mass) 제외 로직
        # 구조적 변형(굽힘, 비틂)은 최소 2개 이상의 블록이 존재해야 정의 가능합니다.
        # 블록이 1개뿐인 컴포넌트는 강체로 간주하고 리포트에서 제외합니다.
        if len(components[comp]) <= 1:
            continue

        log_and_print("=" * 83)
        log_and_print(f"[최대 구조 변형 지표 로컬라이징 리포트] - Body: {comp}")
        log_and_print("-" * 83)
        log_and_print(f"{'Row Index':<9} | {'Max Bending (deg) / Loc':<23} | {'Max Twist (deg) / Loc':<21} | {'Peak Energy (J) / Loc':<21}")
        log_and_print("-" * 83)
        
        j_keys = [k for k in j_data.keys() if isinstance(k, int)]
        for j in sorted(j_keys):
            hist = j_data[j]
            max_b = max(hist['bending'])
            idx_b = hist['bending'].index(max_b)
            loc_b = hist['loc_b'][idx_b]
            
            max_t = max(hist['twist'])
            idx_t = hist['twist'].index(max_t)
            loc_t = hist['loc_t'][idx_t]
            
            max_e = max(hist['energy'])
            idx_e = hist['energy'].index(max_e)
            loc_e = hist['loc_e'][idx_e]
            
            str_b = f"{max_b:>6.2f} (@ {loc_b:<7})"
            str_t = f"{max_t:>6.2f} (@ {loc_t:<7})"
            str_e = f"{max_e:>6.2f} (@ {loc_e:<7})"
            log_and_print(f"y = {j:<5} | {str_b:<23} | {str_t:<21} | {str_e:<21}")
        log_and_print("=" * 83)
        log_and_print("")
        
    # Plotting (only if requested)
    if config.get("plot_results", True):
        # [NEW] txt 파일 출력을 위한 공용 헬퍼 함수
        export_txt_path = os.path.join(output_dir, f'rds-{timestamp}_data.txt')
        f_txt = open(export_txt_path, "w", encoding="utf-8")
        
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as xlImage
            has_xl = True
            xl_wb = Workbook()
            xl_wb.remove(xl_wb.active)
            export_xl_path = os.path.join(output_dir, f'rds-{timestamp}_data.xlsx')
            xl_row_tracker = {}
        except ImportError:
            has_xl = False
            
        def export_plot_data(title, time_arr, data_dict, sheet_name=None, png_path=None):
            f_txt.write(f"#TITLE: {title}\n")
            legends = list(data_dict.keys())
            max_len = max([len(str(l)) for l in legends]) if legends else 0
            num_header_rows = max(1, (max_len + 18) // 19)
            
            for r in range(num_header_rows):
                header = f"#{'Time (s)':<18}" if r == 0 else f"#{'':<18}"
                for leg in legends:
                    header += f"{str(leg)[r*19:(r+1)*19]:<20}"
                f_txt.write(header + "\n")
            
            def fmt(v):
                s = f"{float(v):.8f}"
                if len(s) > 18:
                    s = f"{float(v):.8e}"
                return f"{s:<20}"
                
            for i in range(len(time_arr)):
                line_str = f" {fmt(time_arr[i]).strip():<19}"
                for leg in legends:
                    try: val = data_dict[leg][i]
                    except (IndexError, TypeError): val = 0.0
                    line_str += fmt(val)
                f_txt.write(line_str + "\n")
            f_txt.write("\n")
            
            if has_xl and sheet_name:
                safe_sheet = str(sheet_name)[:31]
                if safe_sheet in xl_wb.sheetnames:
                    ws = xl_wb[safe_sheet]
                else:
                    ws = xl_wb.create_sheet(title=safe_sheet)
                    if png_path and os.path.exists(png_path):
                        try:
                            img = xlImage(png_path)
                            ws.add_image(img, 'A1')
                        except: pass
                
                if safe_sheet not in xl_row_tracker:
                    xl_row_tracker[safe_sheet] = 11
                start_row = xl_row_tracker[safe_sheet]
                
                ws.cell(row=start_row, column=1, value=f"#TITLE: {title}")
                start_row += 1
                
                headers = ['Time (s)'] + legends
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=start_row, column=col_idx, value=str(h))
                    
                for i in range(len(time_arr)):
                    r_idx = start_row + 1 + i
                    ws.cell(row=r_idx, column=1, value=time_arr[i])
                    for col_idx, leg in enumerate(legends, 2):
                        try: val = float(data_dict[leg][i])
                        except: val = 0.0
                        ws.cell(row=r_idx, column=col_idx, value=val)
                xl_row_tracker[safe_sheet] = start_row + 1 + len(time_arr) + 1

        plt.figure(figsize=(10, 5))
        plt.plot(time_history, root_acc_history, label='Internal TV Accel (G)', color='black')
        plt.xlabel('Time (s)')
        plt.ylabel('G-Force')
        plt.title('Internal TV (OpenCell/Chassis) Peak Impact G-Force')
        plt.grid(True)
        plt.legend()
        plt.tight_layout()
        png1 = os.path.join(output_dir, 'rds-impact_gforce.png')
        plt.savefig(png1)
        plt.close()
        export_plot_data('Internal TV (OpenCell/Chassis) Peak Impact G-Force', time_history, {'Internal TV Accel (G)': root_acc_history}, 'Peak Impact G-Force', png1)
        
        # Ground Impact + Air Resistance combined plot (2 subplots)
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 8), sharex=True)
        fig.suptitle('Ground Impact Force & Air Resistance Forces')
        ax1.plot(time_history, ground_impact_hist, label='Ground Normal Force (N)', color='red')
        ax1.set_ylabel('Force (N)')
        ax1.set_title('Ground Impact')
        ax1.grid(True)
        ax1.legend()
        ax2.plot(time_history, air_drag_hist,    label=f'Drag  (Cd={config.get("air_cd_drag",1.05):.2f})', color='blue')
        ax2.plot(time_history, air_viscous_hist, label=f'Viscous (Cd_v={config.get("air_cd_viscous",0.0):.2f})', color='green')
        ax2.plot(time_history, air_squeeze_hist, label=f'Squeeze Film (k={config.get("air_coef_squeeze",1.0):.1f})', color='orange')
        ax2.set_xlabel('Time (s)')
        ax2.set_ylabel('Force (N)')
        ax2.set_title('Air Resistance Components')
        ax2.grid(True)
        ax2.legend()
        plt.tight_layout()
        png2 = os.path.join(output_dir, 'rds-ground_impact.png')
        plt.savefig(png2)
        plt.close()
        
        export_plot_data('Ground Impact Force', time_history, {'Ground Normal Force (N)': ground_impact_hist}, 'Ground Impact', png2)
        export_plot_data('Air Resistance Components', time_history, {
            f'Drag (Cd={config.get("air_cd_drag",1.05):.2f})': air_drag_hist,
            f'Viscous (Cd_v={config.get("air_cd_viscous",0.0):.2f})': air_viscous_hist,
            f'Squeeze Film (k={config.get("air_coef_squeeze",1.0):.1f})': air_squeeze_hist
        }, 'Ground Impact')
        
        # Motion All & Motion Z Setup
        pos_np = np.array(pos_hist) # (steps, 3)
        vel_np = np.array(vel_hist) # (steps, 6) -> [wx, wy, wz, vx, vy, vz]
        acc_np = np.array(acc_hist) # (steps, 6) -> [ax, ay, az, lin_ax, ...]
        
        c_pos_np = np.array(corner_pos_hist) # (steps, 8, 3)
        c_vel_np = np.array(corner_vel_hist) # (steps, 8, 3)
        c_acc_np = np.array(corner_acc_hist) # (steps, 8, 3)
        
        curves = {
            'Center': {'pos': pos_np, 'vel': vel_np[:, 3:6], 'acc': acc_np[:, 3:6]},
            'True_COG': {'pos': np.array(cog_pos_hist), 'vel': np.array(cog_vel_hist), 'acc': np.array(cog_acc_hist)},
            'CORNER_L-B-B': {'pos': c_pos_np[:,0,:], 'vel': c_vel_np[:,0,:], 'acc': c_acc_np[:,0,:]},
            'CORNER_L-B-F': {'pos': c_pos_np[:,1,:], 'vel': c_vel_np[:,1,:], 'acc': c_acc_np[:,1,:]},
            'CORNER_L-T-B': {'pos': c_pos_np[:,2,:], 'vel': c_vel_np[:,2,:], 'acc': c_acc_np[:,2,:]},
            'CORNER_L-T-F': {'pos': c_pos_np[:,3,:], 'vel': c_vel_np[:,3,:], 'acc': c_acc_np[:,3,:]},
            'CORNER_R-B-B': {'pos': c_pos_np[:,4,:], 'vel': c_vel_np[:,4,:], 'acc': c_acc_np[:,4,:]},
            'CORNER_R-B-F': {'pos': c_pos_np[:,5,:], 'vel': c_vel_np[:,5,:], 'acc': c_acc_np[:,5,:]},
            'CORNER_R-T-B': {'pos': c_pos_np[:,6,:], 'vel': c_vel_np[:,6,:], 'acc': c_acc_np[:,6,:]},
            'CORNER_R-T-F': {'pos': c_pos_np[:,7,:], 'vel': c_vel_np[:,7,:], 'acc': c_acc_np[:,7,:]}
        }
        # [REORDERED] Legend 순서를 Corner 우선, Root-relative(Center/COG)를 후순위로 배치함
        ordered_keys = ['CORNER_L-T-F', 'CORNER_R-T-F', 'CORNER_L-T-B', 'CORNER_R-T-B', 
                        'CORNER_L-B-F', 'CORNER_R-B-F', 'CORNER_L-B-B', 'CORNER_R-B-B',
                        'Center', 'True_COG']
        
        # rds-Motion_All.png
        fig, axs = plt.subplots(3, 3, figsize=(18, 12))
        fig.suptitle('Kinematics: Position, Velocity, Acceleration vs Time', fontsize=16)
        labels_row = ['Position (m)', 'Velocity (m/s)', 'Acceleration (m/s^2)']
        labels_col = ['X Axis', 'Y Axis', 'Z Axis']
        all_subplot_dicts = []
        for row in range(3):
            for col in range(3):
                ax = axs[row, col]
                subplot_dict = {}
                for k in ordered_keys:
                    metric = 'pos' if row == 0 else ('vel' if row == 1 else 'acc')
                    # [STYLING] Center와 True_COG는 얇은 점선으로 표시하여 가독성 증대
                    if k in ['Center', 'True_COG']:
                        ax.plot(time_history, curves[k][metric][:, col], label=k, linestyle='--', linewidth=1.0, alpha=0.7)
                    else:
                        ax.plot(time_history, curves[k][metric][:, col], label=k)
                    subplot_dict[k] = curves[k][metric][:, col]
                if row == 2: ax.set_xlabel('Time (s)')
                if col == 0: ax.set_ylabel(labels_row[row])
                if row == 0: ax.set_title(labels_col[col])
                ax.grid(True)
                all_subplot_dicts.append((f"Kinematics - {labels_row[row]} - {labels_col[col]}", subplot_dict))
        axs[0, 2].legend(loc='upper left', bbox_to_anchor=(1.05, 1))
        plt.tight_layout(rect=[0, 0, 0.9, 1])
        png3 = os.path.join(output_dir, 'rds-Motion_All.png')
        plt.savefig(png3)
        plt.close()
        for i, (ttl, s_dict) in enumerate(all_subplot_dicts):
            export_plot_data(ttl, time_history, s_dict, 'Motion_All', png3 if i==0 else None)
        
        # rds-Motion_Z.png
        fig, axs = plt.subplots(1, 3, figsize=(18, 5))
        fig.suptitle('Kinematics (Z Axis Only): Position, Velocity, Acceleration vs Time', fontsize=16)
        titles_z = ['Z Position (m)', 'Z Velocity (m/s)', 'Z Acceleration (m/s^2)']
        z_subplot_dicts = []
        for i, metric in enumerate(['pos', 'vel', 'acc']):
            ax = axs[i]
            subplot_dict = {}
            for k in ordered_keys:
                # [STYLING] Z-Axis 차트에서도 Center/COG를 점선으로 처리
                if k in ['Center', 'True_COG']:
                    ax.plot(time_history, curves[k][metric][:, 2], label=k, linestyle='--', linewidth=1.0, alpha=0.7)
                else:
                    ax.plot(time_history, curves[k][metric][:, 2], label=k)
                subplot_dict[k] = curves[k][metric][:, 2]
            ax.set_xlabel('Time (s)')
            ax.set_ylabel(titles_z[i])
            ax.grid(True)
            z_subplot_dicts.append((f"Kinematics (Z Axis Only) - {titles_z[i]}", subplot_dict))
        axs[2].legend(loc='upper left', bbox_to_anchor=(1.05, 1))
        plt.tight_layout(rect=[0, 0, 0.9, 1])
        png4 = os.path.join(output_dir, 'rds-Motion_Z.png')
        plt.savefig(png4)
        plt.close()
        for i, (ttl, s_dict) in enumerate(z_subplot_dicts):
            export_plot_data(ttl, time_history, s_dict, 'Motion_Z', png4 if i==0 else None)

        # Plot metric per component max over time
        for comp, j_data in metrics.items():
            # [NEW] 블록이 1개인 강체는 시각화 리포트에서도 제외
            if len(components[comp]) <= 1:
                continue

            j_keys = [k for k in j_data.keys() if isinstance(k, int)]
            if len(j_keys) > 0:
                plt.figure(figsize=(10, 5))
                
                # Find maximums across all J rows dynamically
                all_b_hist = np.array([j_data[j]['bending'] for j in j_keys])
                all_t_hist = np.array([j_data[j]['twist'] for j in j_keys])
                
                max_b_time = np.max(all_b_hist, axis=0)
                max_t_time = np.max(all_t_hist, axis=0)
                
                # Find the location string of the absolute global peak to display in legend
                max_b_idx = np.argmax(max_b_time)
                row_idx_b = np.argmax(all_b_hist[:, max_b_idx])
                global_loc_b = j_data[j_keys[row_idx_b]]['loc_b'][max_b_idx]
                
                max_t_idx = np.argmax(max_t_time)
                row_idx_t = np.argmax(all_t_hist[:, max_t_idx])
                global_loc_t = j_data[j_keys[row_idx_t]]['loc_t'][max_t_idx]
                
                plt.plot(time_history, max_b_time, label=f'Max Bending (deg) [{global_loc_b}]')
                plt.plot(time_history, max_t_time, label=f'Max Twisting (deg) [{global_loc_t}]')
                plt.xlabel('Time (s)')
                plt.ylabel('Angle (deg)')
                plt.title(f'{comp} Structural Deformation (Max)')
                plt.grid(True)
                plt.legend()
                plt.tight_layout()
                png_def = os.path.join(output_dir, f'rds-{comp}_deformation.png')
                plt.savefig(png_def)
                plt.close()
                export_plot_data(f'{comp} Structural Deformation (Max)', time_history, {
                    f'Max Bending [{global_loc_b}]': max_b_time,
                    f'Max Twisting [{global_loc_t}]': max_t_time
                }, f'{comp[:15]}_deformation', png_def)
            
            # New: All Blocks Angle Plot
            all_blocks = j_data.get('all_blocks_angle', {})
            block_noms = j_data.get('block_nominals', {})
            
            # [NEW] 블록이 1개뿐인 경우 상대적인 각도 변화 차트가 무의미하므로 제외
            if len(all_blocks) > 0 and len(components[comp]) > 1:
                fig = plt.figure(figsize=(14, 6))
                
                # Create main plot for angle curves
                ax_main = fig.add_axes([0.05, 0.1, 0.65, 0.8])
                num_blocks = len(all_blocks)
                cols_legend = min(4, max(1, num_blocks // 10))
                
                block_dict = {}
                for block_idx, a_hist in all_blocks.items():
                    legend_name = f"{block_idx[0]}-{block_idx[1]}-{block_idx[2]}"
                    ax_main.plot(time_history, a_hist, label=legend_name, linewidth=1.0)
                    block_dict[legend_name] = a_hist
                ax_main.set_xlabel('Time (s)')
                ax_main.set_ylabel('Def. Angle (deg)')
                ax_main.set_title(f'{comp} Block Angle Deformation (Relative to Root)')
                ax_main.grid(True)
                
                # Put Legend outside
                ax_main.legend(loc='upper left', bbox_to_anchor=(1.0, 1.0), ncol=cols_legend, fontsize=6)
                
                # Inset 3D scatter plot for index guide
                # Move slightly further down-right to avoid any potential clash
                ax_inset = fig.add_axes([0.72, 0.05, 0.25, 0.25], projection='3d')
                xs, ys, zs = [], [], []
                labels = []
                for b_idx, nom in block_noms.items():
                    xs.append(nom[0])
                    ys.append(nom[1])
                    zs.append(nom[2])
                    labels.append(f"{b_idx[0]}-{b_idx[1]}-{b_idx[2]}")
                ax_inset.scatter(xs, ys, zs, color='blue', alpha=0.5, s=10)
                
                # Subsample labels if too many to avoid clutter
                step = max(1, len(labels) // 20)
                for i in range(0, len(labels), step):
                    ax_inset.text(xs[i], ys[i], zs[i], labels[i], size=5, zorder=1, color='k')
                
                ax_inset.set_title("i-j-k Reference")
                ax_inset.set_xticks([])
                ax_inset.set_yticks([])
                ax_inset.set_zticks([])
                
                png_def_all = os.path.join(output_dir, f'rds-{comp}_deformation_all.png')
                plt.savefig(png_def_all)
                plt.close()
                export_plot_data(f'{comp} Block Angle Deformation (Relative to Root)', time_history, block_dict, f'{comp[:15]}_def_all', png_def_all)
            
            # [NEW] Total Distortion (SDI/TDI) Plot
            tdi_hist = metrics[comp].get('total_distortion', [])
            if tdi_hist:
                plt.figure(figsize=(10, 5))
                plt.plot(time_history, tdi_hist, color='purple', linewidth=2.0)
                plt.xlabel('Time (s)')
                plt.ylabel('TDI (RMS Angle Deg)')
                plt.title(f'{comp} Total Structural Distortion Index (SDI)')
                plt.grid(True)
                plt.tight_layout()
                png_tdi = os.path.join(output_dir, f'rds-{comp}_total_distortion.png')
                plt.savefig(png_tdi)
                plt.close()
                export_plot_data(f'{comp} Total Structural Distortion Index (SDI)', time_history, {'SDI': tdi_hist}, f'{comp[:15]}_SDI', png_tdi)

            # [NEW] 2x2 Subplot for Corner Plasticity
            corner_hists = metrics[comp].get('corner_hists', {})
            if corner_hists:
                fig, axs = plt.subplots(2, 2, figsize=(14, 10))
                fig.suptitle(f'{comp} Corner Impact Plasticity Analysis (12 Points)', fontsize=16)
                
                titles = ['Instant. Strain', 'Contact Pressure (kPa)', 'Instant. Compression (m)', 'Permanent Deformation (m)']
                keys = ['strain', 'pressure', 'disp', 'plastic']
                
                plot_data_collect = {}
                for idx, key in enumerate(keys):
                    row, col = idx // 2, idx % 2
                    ax = axs[row, col]
                    for cid, hist in corner_hists.items():
                        vals = np.array(hist[key])
                        if key == 'pressure': vals = vals / 1e3 # Pa to kPa
                        ax.plot(time_history, vals, label=hist['name'], alpha=0.7)
                        plot_data_collect[f"{hist['name']}_{key}"] = list(vals)
                    ax.set_title(titles[idx])
                    ax.set_xlabel('Time (s)')
                    ax.set_ylabel(titles[idx])
                    ax.grid(True)
                
                axs[0, 1].legend(loc='upper left', bbox_to_anchor=(1.05, 1.0), fontsize=6)
                plt.tight_layout(rect=[0, 0, 0.85, 0.95])
                png_corner = os.path.join(output_dir, f'rds-{comp}_corner_analysis.png')
                plt.savefig(png_corner)
                plt.close()
                export_plot_data(f'{comp} Corner Plasticity Analysis', time_history, plot_data_collect, f'{comp[:12]}_corner', png_corner)

        f_txt.close()
        if has_xl:
            xl_wb.save(export_xl_path)
            log_and_print(f"  >> Excel saved to: {os.path.basename(export_xl_path)}")
        
    # Copy the XML file used for simulation to the output directory
    if xml_path and os.path.exists(xml_path):
        shutil.copy(xml_path, os.path.join(output_dir, os.path.basename(xml_path)))
        log_and_print(f"  >> Results saved to: {os.path.basename(output_dir)}")
    
    # [Important] Callback 해제 (메모리 릭 및 충돌 방지)
    mujoco.set_mjcb_control(None)
    
    sim_result = DropSimResult(
        config=config,
        metrics=metrics,
        max_g_force=max_g_force,
        time_history=time_history,
        z_hist=z_hist,
        root_acc_history=root_acc_history,
        corner_acc_hist=corner_acc_hist,
        pos_hist=pos_hist,
        vel_hist=vel_hist,
        acc_hist=acc_hist,
        cog_pos_hist=cog_pos_hist,
        cog_vel_hist=cog_vel_hist,
        cog_acc_hist=cog_acc_hist,
        corner_pos_hist=corner_pos_hist,
        corner_vel_hist=corner_vel_hist,
        ground_impact_hist=ground_impact_hist,
        air_drag_hist=air_drag_hist,
        air_viscous_hist=air_viscous_hist,
        air_squeeze_hist=air_squeeze_hist
    )
    result_path = os.path.join(output_dir, f'rds-{timestamp}_result.pkl')
    sim_result.save(result_path)
    log_and_print(f"  >> Python 데이터 오브젝트(DropSimResult) 저장 완료: {os.path.basename(result_path)}")
    
    return sim_result, time_history, z_hist, vel_hist, root_acc_history, corner_acc_hist, max_g_force, metrics

def calculate_required_aux_masses(cfg, target_mass, target_cog, target_moi):
    """
    현재 모델의 관성 특성을 분석하고, 목표로 하는 질량/CoG/MoI를 달성하기 위해 
    다수의 보정 질량(aux_masses) 리스트를 계산하여 반환합니다.
    단일 블록으로는 달성하기 어려운 복합 관성을 맞추기 위해 8개의 대칭된
    포인트 질량체(Point Masses)로 목표치를 완벽히 분배합니다.
    """
    from run_discrete_builder import create_model
    import math
    
    # 1. 기저 모델 관성 획득 (보정 질량 제외)
    temp_cfg = cfg.copy()
    temp_cfg["chassis_aux_masses"] = []  
    
    # create_model 호출하여 현재 상태 측정
    _, m_base, c_base, i_base, _ = create_model("temp_calc.xml", config=temp_cfg, logger=lambda x: None)
    
    m_base = float(m_base)
    c_base = np.array(c_base)
    i_base = np.array(i_base) # [Ixx, Iyy, Izz]

    # 2. 목표값 처리 (None일 경우 Baseline 값 사용)
    target_mass = target_mass if target_mass is not None else m_base
    target_cog  = target_cog  if target_cog  is not None else c_base
    target_moi  = target_moi  if target_moi  is not None else i_base

    # 필요한 총 보정 질량
    m_aux = target_mass - m_base
    
    # 보정할 분량이 거의 없는 경우 빈 리스트 반환
    if abs(m_aux) < 1e-6:
        # 질량은 맞는데 CoG나 MoI가 다를 수도 있으므로 체크
        # 하지만 m_aux=0이면 점질량으로 보정 불가. 최소 질량 요구.
        if (np.allclose(target_cog, c_base, atol=1e-5) and 
            np.allclose(target_moi, i_base, atol=1e-5)):
            return []
        else:
            raise ValueError(f"질량 변화 없이 CoG/MoI만 보정하는 것은 물리적으로 불가능합니다 (최소한의 추가 질량이 필요함).")

    if m_aux < 0:
        raise ValueError(f"목표 질량({target_mass})이 현재 질량({m_base})보다 작습니다. (감량 보정은 지원하지 않음)")

    # 3. 추가 질량체 군집의 목표 CoG 연산
    pos_aux = (np.array(target_cog) * target_mass - m_base * c_base) / m_aux

    # 4. 관성 모멘트 평행축 정리 역산
    def shift_moi(m, i_cg, cg, target_p):
        d = target_p - cg
        ix = i_cg[0] + m * (d[1]**2 + d[2]**2)
        iy = i_cg[1] + m * (d[0]**2 + d[2]**2)
        iz = i_cg[2] + m * (d[0]**2 + d[1]**2)
        return np.array([ix, iy, iz])

    i_base_at_target = shift_moi(m_base, i_base, c_base, target_cog)
    i_aux_req_at_target = np.array(target_moi) - i_base_at_target
    
    # pos_aux 기준에서 가져야할 순수 Inertia 
    i_aux_req_at_own_cog = i_aux_req_at_target - shift_moi(m_aux, [0,0,0], pos_aux, target_cog)
    
    # 이론상 불가능할 경우(음수) 최소 오차를 위해 0으로 제한 (물리적 한계 보정)
    i_res = np.maximum(i_aux_req_at_own_cog, 1e-9)

    # 5. 8개의 질량체로 대칭 분배
    m_each = m_aux / 8.0
    
    # 기하학적 형상(단일 박스)의 구속에 얽매이지 않고, 대칭된 8개의 점으로 텐서 완벽 분해
    # 8-Point Mass 위치 오프셋 해석해
    dx_sq = (i_res[1] + i_res[2] - i_res[0]) / (2.0 * m_aux)
    dy_sq = (i_res[0] + i_res[2] - i_res[1]) / (2.0 * m_aux)
    dz_sq = (i_res[0] + i_res[1] - i_res[2]) / (2.0 * m_aux)
    
    dx = math.sqrt(max(0, dx_sq))
    dy = math.sqrt(max(0, dy_sq))
    dz = math.sqrt(max(0, dz_sq))
    
    s = [0.01, 0.01, 0.01] # 충돌이나 시각을 방해하지 않는 매우 작은 형상 더미
    
    aux_masses = []
    idx = 1
    for sign_x in [-1, 1]:
        for sign_y in [-1, 1]:
            for sign_z in [-1, 1]:
                px = pos_aux[0] + sign_x * dx
                py = pos_aux[1] + sign_y * dy
                pz = pos_aux[2] + sign_z * dz
                
                aux_masses.append({
                    "name": f"AutoB_{idx}",
                    "pos": [float(round(px, 5)), float(round(py, 5)), float(round(pz, 5))],
                    "mass": float(round(m_each, 5)),
                    "size": s
                })
                idx += 1

    return aux_masses

def test_run_case_1():
    # Get basic defaults, override specifically if needed    
    cfg = get_default_config()

    # [GEOMETRY OPTIONS] 포장재 및 제품의 기본 치수 정의 (m 단위)
    cfg["box_w"] = 1.841          # 포장 박스 가로 (m)
    cfg["box_h"] = 1.103          # 포장 박스 세로 (m)
    cfg["box_d"] = 0.170          # 포장 박스 깊이 (m)
    cfg["box_thick"] = 0.008      # 박스 골판지 두께 (m)
    
    cfg["assy_w"] = 1.670         # 내부 제품(TV 등) 가로 (m)
    cfg["assy_h"] = 0.960         # 내부 제품(TV 등) 세로 (m)
    cfg["cush_gap"] = 0.005       # 부품 간 간격 (Clearance)
    
    # 낙하 모드 및 구체적인 방향 설정 (Full Name 사용: front/back, top/bottom, left/right)
    # 형식: [Front/Back]-[Top/Bottom]-[Left/Right] (필요 없는 축은 '--'로 표시)
    # PARCEL 모드 (Standard Mapping):
    # 1: Top, 2: Bottom, 3: Front, 4: Back, 5: Left, 6: Right
    # LTL 모드 (Custom Mapping):
    # 1: Top, 2: Back, 3: Bottom, 4: Front, 5: Right, 6: Left
    #cfg["drop_mode"] = "CUSTOM"
    #cfg["drop_direction"] = "rear-bottom-right" # 예: 전면-아랫면-왼쪽 (꼭짓점 낙하)
    cfg["drop_mode"] = "LTL"
    cfg["drop_direction"] = "Corner 2-3-5" # 예: 전면-아랫면-왼쪽 (꼭짓점 낙하)
    cfg["include_paperbox"] = False # 로컬 테스트용 오버라이드
    cfg["drop_height"] = 0.5    
    cfg["plot_results"] = True    
    cfg["only_generate_xml"] = False     # [NEW] True 설정 시 시뮬레이션을 돌리지 않고 XML 모델 생성 및 저장만 수행

    # [COMPONENTS OPTIONS] 부품별 해상도(div) 및 결합 방식(use_weld) 설정
    # use_weld=False 설정 시, 해당 부품은 내부 구속조건이 없는 '단일 강체'로 취급되어 연산 속도가 비약적으로 향상됩니다.
    cfg["chassis_div"]      = [3, 3, 1]    # 샤시 분할 수
    cfg["chassis_use_weld"] = True        # 샤시는 강체로 취급 (속도 향상)
    
    cfg["oc_div"]           = [3, 3, 1]    # 오픈셀(패널) 분할 수
    cfg["oc_use_weld"]      = True         # 오픈셀은 강체로 취급 (속도 향상)
    
    cfg["occ_div"]          = [3, 3, 1]    # 테이프(Cohesive) 분할 수
    cfg["occ_use_weld"]     = True         # 테이프는 유연성 유지를 위해 Weld 사용
    
    cfg["cush_div"]         = [5, 5, 3]    # 쿠션 분할 수
    cfg["cush_use_weld"]    = True         # 쿠션은 변형이 중요하므로 Weld 사용
    
    cfg["box_div"]          = [5, 5, 2]    # 외곽 박스 분할 수 (성능을 위해 낮춤)
    cfg["box_use_weld"]     = False        # 박스는 강체로 취급
    
    # [NEW] 물리 파라미터 고도화 (Weld vs Contact 분리 및 엣지 특화)
    # 1. 쿠션 (Cushion)
    cfg["cush_weld_solref_timec"] = 0.004      # 쿠션 전체 구조적 벤딩 강성 (Stiffness)
    cfg["cush_weld_solref_dampr"]  = 1.0      # 쿠션 전체 구조적 감쇠 (Damping)
    
    # [NEW] 코너 쿠션 전용 Weld 강성 (구조 강성과 분리하여 코너 압축 특성 특화)
    cfg["cush_weld_corner_solref_timec"] = 0.02 # 코너 부위 전용 stiffness (timeconst)
    cfg["cush_weld_corner_solref_damprr"] = 1.0  # 코너 부위 전용 damping (dampratio)
    
    cfg["cush_contact_solref"]    = "0.01 0.8" # 일반 접촉 (Center)
    cfg["cush_contact_solimp"]    = "0.1 0.95 0.005 0.5 2" 
    cfg["cush_corner_solref"]     = "0.01 0.8" # 엣지/모너리 전용 접촉 (Edge/Corner)
    cfg["cush_corner_solimp"]     = "0.1 0.95 0.005 0.5 2"
    # [SOLVER OPTIONS] MuJoCo 물리 엔진 및 솔버 관련 상세 설정 (sol_*)
    #cfg["cush_solref_timec"] = 0.01 # 쿠션 timeconst
    #cfg["cush_solref_dampr"]  = 0.1 # 쿠션 dampratio

    # 2. 테이프/접착제 (Tape)
    #cfg["tape_weld_solref"] = "0.01 1.0"
    #cfg["tape_solref"]      = "0.05 1.0"
    
    # 3. 오픈셀/패널 (Cell/TV)
    #cfg["cell_weld_solref"] = "0.005 1.0"
    #cfg["cell_solref"]      = "0.01 1.0"
    
    # [NEW] 영구 변형 테스트 활성화
    cfg["enable_plasticity"] = True
    cfg["plasticity_ratio"] = 1.0
    cfg["cush_yield_strain"] = 0.01   # 1% 변형 시 소성 변형 가능성 (알고리즘 2번용)
    cfg["cush_yield_pressure"] = 80.0  # [NEW] 최소 0.5kPa 이상의 압력이 있어야 소성 변형 발생
    cfg["plasticity_algorithm"] = 2  # 1: Pressure/Penetration, 2: Strain(Neighbor Distance)
    
    cfg["mass_paper"] = 4.0
    cfg["mass_cushion"] = 2.0
    cfg["mass_oc"] = 5.0
    cfg["mass_occ"] = 0.1
    cfg["mass_chassis"] = 10.0
        
    # [RECOMMENDED] 딱딱한 바닥과의 충돌 시 관통 방지를 위한 설정
    # ground_solref: [timeconst, dampratio] -> timeconst가 작을수록 딱딱함. (0.002 미만은 비권장)
    cfg["ground_solref_timec"] = 0.001  # 0.01에서 0.004로 강화 (더 딱딱한 바닥)
    cfg["ground_solref_dampr"]  = 0.0001    # Critical Damping
    cfg["ground_friction"]     = 0.1   # 바닥 마찰계수

    # ground_solimp: [dmin, dmax, width, midpoint, power]     
    cfg["ground_solimp"] = "0.1 0.95 0.001 0.5 2"
    
    # [SIMULATION OPTIONS] MuJoCo 물리 엔진 및 솔버 관련 상세 설정 (sim_*)
    cfg["sim_integrator"] = "implicitfast" # 통합기 (Euler, RK4, implicit, implicitfast 등)
    cfg['sim_duration'] = 2.1
    cfg["sim_timestep"]   = 0.0013         # 시뮬레이션 타임스텝 (s)
    cfg["sim_iterations"] = 50             # 솔버 최대 반복 횟수 (1~200)
    cfg["sim_noslip_iterations"] = 0       # 마찰 고정(미끄러짐 방지)을 위한 추가 반복 횟수
    cfg["sim_tolerance"]  = 1e-5           # 솔버 수렴 오차 허용치 (속도를 위해 약간 완화)
    cfg["sim_gravity"]    = [0, 0, -9.81]  # 중력 가속도 [x, y, z]
    cfg["sim_nthread"]    = 4              # [NEW] 멀티코어 사용 (코어 수에 맞춰 조절)
    cfg["sim_impratio"]   = 5.0             # 기본 1.0. 값을 5~10 정도로 크게 주면 관통을 강력하게 봉쇄합니다.

    

    # air fluidic force
    cfg["air_density"]      = 1.225     # 공기 밀도 (kg/m^3, 20도 1atm)
    cfg["air_viscosity"]    = 1.81e-5   # 공기 동점성계수 (Pa.s)
    cfg["air_cd_drag"]      = 1.05      # Blunt drag 계수 (박스 형태 기준 1.0~1.2)
    cfg["air_cd_viscous"]   = 0.01      # Slender(점성) drag 계수 (박스는 보통 0)
    cfg["air_coef_squeeze"] = 0.2       # Squeeze Film 효과 강도 배율 (0=비활성화)
    cfg["air_squeeze_hmax"] = 0.20      # Squeeze Film 활성화 최대 높이 (m)
    cfg["air_squeeze_hmin"] = 0.005      # Squeeze Film 최소 높이 (분모 안전값, m)
    cfg["enable_air_drag"]    = True   # MuJoCo 빌트인 Drag/Viscous 활성화 여부
    cfg["enable_air_squeeze"] = False   # 수동 Squeeze Film 활성화 여부

    # [STEP 1] 보정 전 설계 원안 검토 (Baseline Review)
    print_inertia_report(cfg, title="Baseline Inertia Report (Pre-Correction)", logger=print)
    
    # [STEP 2] 자동 질량 보정 적용 (Target 맞춤)
    target_mass = 25.0
    target_cog = [0.0, -0.001, -0.001]
    target_cog = None
    #target_moi = [3.0, 8.0, 10.0]  # [Ixx, Iyy, Izz]
    target_moi = None

    try:
        # 예시: CoG만 보정하고 싶다면 target_mass=None, target_moi=None 입력 가능
        aux_items = calculate_required_aux_masses(cfg, target_mass, target_cog, target_moi)
        cfg["chassis_aux_masses"] = aux_items
        print(f"\n>> [Inertia Balancer] Target 기반 보정 질량 (8-Point 다중 질량체) 생성 완료:")
        print(f"      - 생성된 개체 수: {len(aux_items)}EA, 총 추가 질량: {sum(x['mass'] for x in aux_items):.4f}kg")
        if aux_items:
            for aux in aux_items[:2]:
                print(f"      - 예시: {aux['name']} / Pos: {aux['pos']} / Mass: {aux['mass']}kg")
            print("      ... 등격(Symmetric) 배치됨")
        else:
            print("      - 보정이 필요하지 않은 상태입니다 (Baseline 유지).")
    except Exception as e:
        print(f"\n>> [Inertia Balancer] 보정 계산 실패 (기본값 사용): {e}")
        cfg["chassis_aux_masses"] = []
    
    # [STEP 3] 실행 전 최종 확인 (선택 사항)
    print("\n>> 보정이 적용된 최종 모델 구성을 확인합니다...")
    print_inertia_report(cfg, title="Final Balanced Inertia Report", logger=print)

    # [STEP 4] 실행 시 뷰어 사용 여부 확인 후 시뮬레이션 시작
    cfg["use_viewer"] = True # 로컬 테스트 시 뷰어 비활성화
    run_simulation(cfg)


if __name__ == "__main__":
    test_run_case_1()

'''
1. 재료별 추천 
solref 값 가이드
- 재료 유형	추천 timeconst (강성)	추천 dampratio (감쇠)	특징 및 설명
- Rigid (금속, 딱딱한 플라스틱)	0.002 ~ 0.005	1.0 ~ 2.0	변형이 거의 없는 소자. 타임스텝의 2배(0.002)가 수치적 한계치입니다.
- Semi-Rigid (골판지, 강화 플라스틱)	0.01 ~ 0.02	1.0	약간의 탄성이 느껴지지만 기본적으로 형태를 유지해야 하는 경우입니다.
- Adhesive/Tape (점착제, 테이프)	0.005 ~ 0.02	1.0 ~ 1.5	현재 프로젝트의 Cohesive 층에 해당합니다. 결합력이 강할수록 낮은 값을 씁니다.
- Foam/Cushion (스티로폼, EPP)	0.02 ~ 0.1	0.5 ~ 1.0	충격을 흡수하며 눌려야 하는 재질입니다. 값이 클수록 말랑해집니다.
- Soft Rubber (부드러운 고무)	0.05 ~ 0.2	0.1 ~ 0.5	복원력이 강하지만 아주 부드럽게 눌리는 재질입니다.

2. 파라미터별 세부 조정 팁
Timeconst (첫 번째 값)
- 의미: 스프링-댐퍼 시스템이 평형 상태로 돌아오는 데 걸리는 시간 상수입니다.
- Rule of Thumb: 시뮬레이션 타임스텝(dt)의 최소 2배 이상을 권장합니다.
- dt=0.001일 때 0.002가 물리적 강성의 최대치입니다. 이보다 낮으면 수치적 폭발(Explosion)이 일어날 확률이 급격히 높아집니다.
- 조정: 물체가 너무 잘 뚫고 지나가면(Penetration) 이 값을 낮추고, 시뮬레이션이 불안정하게 떨리면 이 값을 높여야 합니다.

Dampratio (두 번째 값)
- 의미: 감쇠비입니다. 1.0이 임계 감쇠(Critical Damping)입니다.
- 1.0: 에너지를 가장 안정적으로 소산시키며 진동 없이 멈춥니다. (대부분의 조립체 권장)
- > 1.0 (Over-damping): 충격 시 튕겨나가는 현상을 억제하고 싶을 때 사용합니다. 낙하 시험에서 바닥과의 접촉 등에 유리합니다.
- < 1.0 (Under-damping): 물체가 통통 튀는 탄성 효과를 주고 싶을 때 사용합니다.

3. 현재 프로젝트(WHToolsBox) 적용 예시
사용자께서 방금 수정하신 값들을 기준으로 분석해 보면 다음과 같습니다.

- Chassis / TV Panel (0.01): 상당히 단단한 구조체로 적절한 설정입니다.
- Tape (0.05): 이전(0.005)보다 10배 더 말랑하게 수정하셨습니다. 이 경우 285G 같은 고충격에서는 테이프가 고무줄처럼 늘어나면서 부품 분리가 일어날 가능성이 매우 높습니다. (분리 문제가 다시 발생한다면 0.01 이하로 낮추는 것을 권장합니다.)
- Cushion (0.1): 아주 부드러운 완충재 설정입니다. 충격 흡수량은 많아지지만 패널이 쿠션을 깊게 파고들 수 있습니다.


1. i-j-k reference 그림과 legend가 겹치는 문제가 있는데 해결할 수 있을까?
2. 결과를 저장하는 python 변수들이 있다면, 이를 저장해놓았다가 다음에 읽어서 처리할 수 있으면 좋겠다. 이를 관장할 수 있는 데이터 클래스를 만들어 파일 저장, 읽기 등과 함게 데이터 처리, 가공, txt 및 excel 저장 등도 총괄 할 수 있도록 하자.

그리고 시뮬레이션 후에 rds 폴더에 해당 결과도 저장해놓고, 다음에 읽어서 데이터를 다뤌 수 있도록 해달라. 

우리는 나중에 주어진 시험 결과(예를 들면, 상자 모서리의 시간에 대한 좌표 변화)와 매칭되는 시뮬레이션 결과를 얻기 위해서 파라미터들을 자동으로 찾아 조정하는 기능을 구현할 게획이다. 이 계획도 참고하라.

'''




'''
ISTA 6-Amazon.com (SIOC) 규격에 따른 낙하 시험에서는 제품의 각 면(Face), 능(Edge), 꼭짓점(Corner)을 숫자로 정의하여 관리합니다. 
사용자가 drop_direction에 "2-3-5"와 같이 숫자를 입력했을 때 이를 정확히 해석하기 위해서는 우선 PARCEL과 LTL 모드에 따른 넘버링 체계를 이해해야 합니다.

1. ISTA 6A 면(Face) 넘버링의 기본 원칙
박스형 포장재의 6개 면은 보통 다음과 같이 정의됩니다:

1번 면 (Face 1): 상면 (Top)
2번 면 (Face 2): 하면 (Bottom / Base) - 가장 중요
3번 면 (Face 3): 전면 (Front / Display side) - TV의 경우 화면부
4번 면 (Face 4): 후면 (Rear / Back)
5번 면 (Face 5): 좌측면 (Left)
6번 면 (Face 6): 우측면 (Right)
NOTE
넘버링 기준 잡기 제품을 정상적인 적재 상태(정면이 나를 향하도록)로 놓았을 때, 천장이 1번, 바닥이 2번입니다. 
그 상태에서 내 눈앞에 보이는 정면이 3번, 그 반대편이 4번이며, 나의 왼손 쪽이 5번, 오른손 쪽이 6번이 됩니다.

2. PARCEL vs LTL: 모드별 넘버링의 차이
ISTA 6A 규격 내에서도 배송 형태(SIOC Type)에 따라 강조되는 지점과 넘버링의 의미가 미세하게 달라집니다.

2.1. PARCEL (Small/Medium - Type A/B)
특징: 택배 기사님이 직접 손으로 들거나 컨베이어에 태우는 개별 포장 방식입니다.
테스트: 모든 방향(6면, 12능, 8코너)에 대한 낙하 강도가 골고루 중요합니다.
넘버링: 표준 1~6번 체계를 엄격히 따르며, 주로 코너 2-3-5 (바닥-전면-왼쪽이 만나는 꼭짓점) 낙하부터 시험을 시작하는 경우가 많습니다.

2.2. LTL (Large/Palletized - Type C/D/E/F)
특징: 지게차나 팰럿 잭을 사용하여 운반하는 대형 화물(TV, 세탁기 등)입니다.
테스트: 바닥면(2번 면) 기준의 수평 낙하와, 기울여서 떨어뜨리는 **회전 낙하(Rotational Drop)**가 주를 이룹니다.
넘버링 차이:
LTL 모드에서는 **물리적 바닥(Base)**이 항상 2번 면이 되어야 합니다.
특히 TV와 같이 슬림한 대형 제품은 포장 상태에 따라 '전면(3)'과 '측면(5/6)'의 정의가 시험 장비 세팅(Clamping 등)에 따라 매우 구체적으로 지정됩니다.
면적이 넓은 면이 1/2번이 될지, 3/4번이 될지에 따라 낙하 충격 에너지가 분산되는 양상이 다르므로 주의가 필요합니다.

3. 숫자 조합을 통한 낙하 위치 정의
숫자를 조합하면 drop_direction을 더욱 명확하게 표현할 수 있습니다.

면 낙하 (Face Drop): 숫자 1개 (3 -> Front 면 낙하)
능 낙하 (Edge Drop): 숫자 2개 (2-3 -> 바닥과 전면이 만나는 모서리 낙하)
꼭짓점 낙하 (Corner Drop): 숫자 3개 (2-3-5 -> 바닥/전면/좌측이 만나는 꼭짓점 낙하)


Parcel 시험
- 1=Top, 2=Bottom, 3=Front, 4=Back, 5=Left, 6=Right.

 LTL 시험 
 - 팔레트 적재 방향을 기준으로 면 번호를 지정




'''
